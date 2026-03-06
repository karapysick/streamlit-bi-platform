# pages_/page_5_approval_status.py

import os
import tempfile
import traceback
from datetime import datetime, timedelta
from io import BytesIO

import matplotlib.pyplot as plt
import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from approval_status import everyday_1 as everyday
from approval_status.dymanics_full import ReportGenerator
from approval_status.full_report import build_full_report_streamlit


APPR_EXCEL_BYTES = "appr_excel_bytes"
APPR_EXCEL_NAME = "appr_excel_name"
APPR_ISSUES_DF = "appr_issues_df"
APPR_FILTERED_COUNT = "appr_filtered_count"
APPR_LAST_RUN = "appr_last_run"
APPR_SHOW_ISSUES = "appr_show_issues"
DB_EXCEL_BYTES = "db_excel_bytes"
APPR_GZ_YEAR_USED = "appr_gz_year_used"
APPR_DF_BASE = "appr_df_base"


def _init_session_state() -> None:
    defaults = {
        APPR_EXCEL_BYTES: None,
        APPR_EXCEL_NAME: None,
        APPR_ISSUES_DF: None,
        APPR_FILTERED_COUNT: None,
        APPR_LAST_RUN: None,
        APPR_SHOW_ISSUES: False,
        DB_EXCEL_BYTES: None,
        APPR_GZ_YEAR_USED: None,
        APPR_DF_BASE: None,
    }

    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value


def _reset_session_state() -> None:
    st.session_state[APPR_EXCEL_BYTES] = None
    st.session_state[APPR_EXCEL_NAME] = None
    st.session_state[APPR_ISSUES_DF] = None
    st.session_state[APPR_FILTERED_COUNT] = None
    st.session_state[APPR_LAST_RUN] = None
    st.session_state[APPR_SHOW_ISSUES] = False
    st.session_state[DB_EXCEL_BYTES] = None
    st.session_state[APPR_GZ_YEAR_USED] = None
    st.session_state[APPR_DF_BASE] = None


def _issues_to_excel_buffer(issues_df: pd.DataFrame) -> BytesIO:
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        issues_df.to_excel(writer, index=False, sheet_name="Проблемные_заказы")
    buffer.seek(0)
    return buffer


def show_approval_status(df: pd.DataFrame) -> None:
    st.header("✅ Отчёт: Статус утверждения")

    _init_session_state()
    df_base = df

    st.sidebar.subheader("Параметры отчёта")

    year_mode = st.sidebar.radio(
        "Год ГЗ",
        ["Авто", "Выбрать"],
        horizontal=True,
        key="appr_year_mode",
    )

    gz_year = None
    if year_mode == "Выбрать":
        gz_year = st.sidebar.selectbox(
            "Год",
            [2024, 2025, 2026, 2027, 2028, 2029, 2030],
            index=1,
            key="appr_year",
        )
        if gz_year >= 2000:
            gz_year = gz_year % 100

    db_bytes = st.session_state.get(DB_EXCEL_BYTES)

    if db_bytes is None:
        st.sidebar.warning("Сначала загрузи БД в начале приложения.")
        db_file = st.sidebar.file_uploader(
            "Excel БД (фолбэк)",
            type=["xlsx", "xls"],
            key="appr_db",
        )
        if db_file is not None:
            db_bytes = db_file.getvalue()
            st.session_state[DB_EXCEL_BYTES] = db_bytes
    else:
        st.sidebar.success("БД уже загружена ✅")

    if st.session_state[APPR_LAST_RUN]:
        st.caption(f"Последняя генерация: {st.session_state[APPR_LAST_RUN]}")

    if st.button("🧾 Сформировать отчёт"):
        try:
            if db_bytes is None:
                st.error("Загрузи Excel БД в сайдбаре.")
                st.stop()

            st.session_state[APPR_SHOW_ISSUES] = False

            db_bytes = st.session_state.get(DB_EXCEL_BYTES)
            if not db_bytes:
                st.error("Сначала загрузи Excel-файл в левой панели.")
                st.stop()

            st.session_state[DB_EXCEL_BYTES] = db_bytes
            st.session_state[APPR_GZ_YEAR_USED] = gz_year
            st.session_state[APPR_DF_BASE] = df_base

            with st.spinner("Формируем отчёт..."):
                excel_bytes, filename, issues_df, filtered_count = build_full_report_streamlit(
                    db_excel_bytes=db_bytes,
                    gz_year=gz_year,
                )

            st.session_state[APPR_EXCEL_BYTES] = excel_bytes
            st.session_state[APPR_EXCEL_NAME] = filename
            st.session_state[APPR_ISSUES_DF] = (
                issues_df if isinstance(issues_df, pd.DataFrame) else pd.DataFrame()
            )
            st.session_state[APPR_FILTERED_COUNT] = (
                int(filtered_count) if filtered_count is not None else None
            )
            st.session_state[APPR_LAST_RUN] = datetime.now().strftime("%d.%m.%Y %H:%M:%S")

            st.success("Отчёт сформирован ✅")

        except Exception as exc:
            st.error(f"Ошибка: {exc}")
            st.code(traceback.format_exc())

    if st.session_state[APPR_FILTERED_COUNT] is not None:
        st.metric("Строк после фильтрации", st.session_state[APPR_FILTERED_COUNT])

    if st.session_state[APPR_EXCEL_BYTES] is not None:
        st.download_button(
            label="📥 Скачать отчёт",
            data=st.session_state[APPR_EXCEL_BYTES],
            file_name=st.session_state[APPR_EXCEL_NAME] or "report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        if st.button("🔍 Показать проблемные заказы"):
            st.session_state[APPR_SHOW_ISSUES] = True

        if st.session_state[APPR_SHOW_ISSUES]:
            issues_df = st.session_state[APPR_ISSUES_DF]

            if isinstance(issues_df, pd.DataFrame) and not issues_df.empty:
                st.subheader("Проблемные заказы")
                st.dataframe(issues_df, width="stretch")

                issues_buffer = _issues_to_excel_buffer(issues_df)

                st.download_button(
                    "📥 Скачать проблемные заказы (xlsx)",
                    data=issues_buffer,
                    file_name="problem_orders.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            else:
                st.success("Проблемных записей нет ✅")

    if st.session_state[APPR_EXCEL_BYTES] is not None:
        with st.expander("👀 Превью отчёта (лист «Шт.»)", expanded=True):
            workbook = load_workbook(
                BytesIO(st.session_state[APPR_EXCEL_BYTES]),
                data_only=True,
            )
            sheet_name = "Шт." if "Шт." in workbook.sheetnames else workbook.sheetnames[0]
            worksheet = workbook[sheet_name]

            max_rows = st.slider("Сколько строк показать", 30, 400, 160, 10)
            max_cols = st.slider("Сколько колонок показать", 10, 80, 45, 1)

            preview_rows = []
            for row in worksheet.iter_rows(
                min_row=1,
                max_row=min(worksheet.max_row, max_rows),
                max_col=min(worksheet.max_column, max_cols),
                values_only=True,
            ):
                preview_rows.append(list(row))

            preview_df = pd.DataFrame(preview_rows)

            while len(preview_df) > 0 and preview_df.tail(1).isna().all(axis=1).iloc[0]:
                preview_df = preview_df.iloc[:-1]

            st.dataframe(preview_df, width="stretch")

    if st.session_state[DB_EXCEL_BYTES] is not None:
        with st.expander("📈 Превью графика (как в GitHub-отчёте)", expanded=True):
            try:
                db_bytes = st.session_state[DB_EXCEL_BYTES]
                gz_year_used = st.session_state.get(APPR_GZ_YEAR_USED)

                temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
                temp_file.write(db_bytes)
                temp_file.close()

                try:
                    generator = ReportGenerator(temp_file.name)

                    gz_year_real = (
                        int(gz_year_used)
                        if gz_year_used is not None
                        else generator.infer_gz_year()
                    )
                    github_df = generator.filter_by_gz_year(
                        generator.selection_filtered,
                        gz_year_real,
                    )

                    today = datetime.today().replace(
                        hour=0,
                        minute=0,
                        second=0,
                        microsecond=0,
                    )
                    weekday = today.weekday()
                    end_date = (
                        today - timedelta(days=1)
                        if weekday == 0
                        else today - timedelta(days=weekday) + timedelta(days=6)
                    )
                    start_date = end_date - timedelta(days=20)
                    date_range = pd.date_range(start=start_date, end=end_date, freq="D")

                    months_ru = [
                        "янв", "фев", "мар", "апр", "май", "июн",
                        "июл", "авг", "сен", "окт", "ноя", "дек",
                    ]
                    periods = [f"{day.day} {months_ru[day.month - 1]}" for day in date_range]

                    workday_mask = [day.weekday() < 5 for day in date_range]
                    yesterday = today - timedelta(days=1)
                    valid_indexes = [i for i, day in enumerate(date_range) if day <= yesterday]
                    last_index = max(valid_indexes) if valid_indexes else len(date_range) - 1

                    periods_chart = periods[: last_index + 1]
                    date_range_chart = date_range[: last_index + 1]
                    workday_mask_chart = workday_mask[: last_index + 1]

                    x_labels = [label for label, is_workday in zip(periods_chart, workday_mask_chart) if is_workday]
                    work_dates = [day for day, is_workday in zip(date_range_chart, workday_mask_chart) if is_workday]

                    series = []
                    for operation in generator.operation_order:
                        column_name = generator.operation_map[operation]
                        values = []

                        for day in work_dates:
                            normalized_day = pd.Timestamp(day).normalize()
                            count = int((github_df[column_name].dt.normalize() == normalized_day).sum())
                            values.append(count)

                        label = generator.chart_series_titles.get(operation, operation)
                        series.append((label, values))

                    fig, ax = plt.subplots()
                    x_positions = list(range(len(x_labels)))

                    for label, values in series:
                        ax.plot(x_positions, values, linewidth=2, label=label)

                    ax.set_xticks(x_positions)
                    ax.set_xticklabels(x_labels, rotation=0, fontsize=8)
                    ax.grid(True, axis="y", linewidth=0.8)
                    ax.grid(False, axis="x")
                    ax.legend(loc="upper left", fontsize=8)
                    ax.set_title(f"Динамика операций по дням (ГЗ-{gz_year_real})")

                    st.pyplot(fig, clear_figure=True)

                finally:
                    try:
                        os.remove(temp_file.name)
                    except Exception:
                        pass

            except Exception:
                st.error("Не удалось построить превью графика GitHub.")
                st.code(traceback.format_exc())
    else:
        st.info("Загрузи БД и нажми «Сформировать отчёт», чтобы появилось превью GitHub-графика.")

    if st.session_state[APPR_FILTERED_COUNT] is not None:
        with st.expander("📋 Сводки статусов (Шт / Га) + 🥧 диаграмма", expanded=True):
            try:
                gz_year_used = st.session_state.get(APPR_GZ_YEAR_USED)

                df_for_calc = st.session_state.get(APPR_DF_BASE)
                if df_for_calc is None:
                    df_for_calc = df

                prepared_df, _ = everyday.prepare_filtered_df(df_for_calc, gz_year_used)
                approved_mask, rejected_mask, review_mask = everyday.build_masks(prepared_df)

                approved_count = int(approved_mask.sum())
                rejected_count = int(rejected_mask.sum())
                review_count = int(review_mask.sum())

                qty_table = pd.DataFrame(
                    {
                        "Статус": ["Утверждено", "Отклонено", "На рассмотрении", "Итого"],
                        "Шт": [
                            approved_count,
                            rejected_count,
                            review_count,
                            approved_count + rejected_count + review_count,
                        ],
                    }
                )

                st.subheader("Сводка по статусам (Шт)")
                st.dataframe(qty_table, width="stretch", hide_index=True)

                hectares_column = "Сумма Объем заказа, га"
                if hectares_column in prepared_df.columns:
                    hectares_series = pd.to_numeric(
                        prepared_df[hectares_column],
                        errors="coerce",
                    )

                    approved_hectares = float(hectares_series[approved_mask].sum(skipna=True))
                    rejected_hectares = float(hectares_series[rejected_mask].sum(skipna=True))
                    review_hectares = float(hectares_series[review_mask].sum(skipna=True))

                    hectares_table = pd.DataFrame(
                        {
                            "Статус": ["Утверждено", "Отклонено", "На рассмотрении", "Итого"],
                            "Га": [
                                approved_hectares,
                                rejected_hectares,
                                review_hectares,
                                approved_hectares + rejected_hectares + review_hectares,
                            ],
                        }
                    )

                    st.subheader("Сводка по статусам (Га)")
                    st.dataframe(hectares_table, width="stretch", hide_index=True)
                else:
                    st.warning(
                        f"Колонка «{hectares_column}» не найдена — сводка по Га недоступна."
                    )

                total_count = approved_count + rejected_count + review_count
                st.subheader("Круговая диаграмма статусов (превью)")

                if total_count == 0:
                    st.warning("По выбранным фильтрам сумма статусов = 0 — диаграмма не строится.")
                else:
                    fig, ax = plt.subplots()
                    ax.pie(
                        [approved_count, rejected_count, review_count],
                        labels=["Утверждено", "Отклонено", "На рассмотрении"],
                        autopct="%1.1f%%",
                    )
                    ax.axis("equal")
                    st.pyplot(fig, clear_figure=True)

            except Exception:
                st.error("Ошибка при построении сводок/диаграммы.")
                st.code(traceback.format_exc())

    if st.button("🧹 Сбросить результат"):
        _reset_session_state()
        st.rerun()