# pages_/page_6_dynamics.py

import datetime as dt
import traceback
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from dynamics_streamlit import build_dynamics_report_streamlit


DYN_EXCEL_BYTES = "dyn_excel_bytes"
DYN_FILENAME = "dyn_filename"
DYN_LAST_RUN = "dyn_last_run"

DB_EXCEL_BYTES = "db_excel_bytes"

DEFAULT_FILENAME = "dynamics.xlsx"


def _init_session_state() -> None:
    defaults = {
        DYN_EXCEL_BYTES: None,
        DYN_FILENAME: DEFAULT_FILENAME,
        DYN_LAST_RUN: None,
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value


def _smart_default_sheets(sheet_names: list[str]) -> list[str]:
    """Select default sheets for preview based on common priority keywords."""
    if not sheet_names:
        return []

    priority_keywords = [
        "шт",
        "ga",
        "га",
        "итого",
        "свод",
        "summary",
        "динамик",
        "общ",
        "total",
    ]

    def score(sheet_name: str) -> int:
        normalized_name = sheet_name.lower()
        total_score = 0
        for index, keyword in enumerate(priority_keywords):
            if keyword in normalized_name:
                total_score += 100 - index
        return total_score

    ranked_sheet_names = sorted(sheet_names, key=score, reverse=True)

    selected_sheet_names: list[str] = []
    for sheet_name in ranked_sheet_names:
        if sheet_name not in selected_sheet_names:
            selected_sheet_names.append(sheet_name)
        if len(selected_sheet_names) >= 3:
            break

    if all(score(name) == 0 for name in selected_sheet_names):
        return sheet_names[:2] if len(sheet_names) >= 2 else sheet_names

    return selected_sheet_names


def _filter_preview_sheets(sheet_names: list[str]) -> list[str]:
    trash_keywords = ("print", "печать", "tmp", "temp", "chart", "график")
    filtered_sheet_names: list[str] = []

    for sheet_name in sheet_names:
        normalized_name = sheet_name.lower()

        if normalized_name.startswith("__"):
            continue

        if any(keyword in normalized_name for keyword in trash_keywords):
            if normalized_name not in ("шт.", "га.", "шт", "га"):
                continue

        filtered_sheet_names.append(sheet_name)

    return filtered_sheet_names


def show_dynamics(df: pd.DataFrame) -> None:
    del df  # data is loaded centrally; this page uses shared bytes from session state

    st.header("📈 Отчёт: Динамика")
    _init_session_state()

    st.sidebar.subheader("Параметры отчёта")

    report_kind = st.sidebar.radio(
        "Тип отчёта",
        ["Ежедневный", "Недельный", "Месячный"],
        index=0,
        key="dyn_kind",
    )

    gz_year = st.sidebar.selectbox(
        "Год ГЗ (2 цифры)",
        [24, 25, 26, 27, 28, 29, 30],
        index=1,
        key="dyn_gz_year",
    )

    plan_count = st.sidebar.number_input(
        "План (шт)",
        min_value=1,
        value=100,
        step=1,
        key="dyn_plan_count",
    )

    plan_area = st.sidebar.number_input(
        "План (га)",
        min_value=0.1,
        value=10.0,
        step=0.1,
        key="dyn_plan_area",
    )

    split_by_ogx = st.sidebar.checkbox(
        "Разбивка по типу ОГХ",
        value=False,
        key="dyn_split_ogx",
    )

    split_by_itp = st.sidebar.checkbox(
        "Разбивка по ИТП",
        value=False,
        key="dyn_split_itp",
    )

    if st.button("🧾 Сформировать отчёт"):
        try:
            db_bytes = st.session_state.get(DB_EXCEL_BYTES)
            if not db_bytes:
                st.error("Сначала загрузите Excel-файл в начале приложения (в левой панели).")
                st.stop()

            with st.spinner("Формируем отчёт..."):
                excel_bytes, filename = build_dynamics_report_streamlit(
                    db_excel_bytes=db_bytes,
                    report_kind=report_kind,
                    gz_year=int(gz_year),
                    plan_count=int(plan_count),
                    plan_area=float(plan_area),
                    split_by_ogx=bool(split_by_ogx),
                    split_by_itp=bool(split_by_itp),
                )

            st.session_state[DYN_EXCEL_BYTES] = excel_bytes
            st.session_state[DYN_FILENAME] = filename
            st.session_state[DYN_LAST_RUN] = dt.datetime.now().strftime("%d.%m.%Y %H:%M:%S")

            st.success("Отчёт сформирован ✅")

        except Exception:
            st.error("Ошибка при формировании отчёта.")
            st.code(traceback.format_exc())

    if st.session_state[DYN_EXCEL_BYTES] is not None:
        st.download_button(
            "📥 Скачать отчёт",
            data=st.session_state[DYN_EXCEL_BYTES],
            file_name=st.session_state[DYN_FILENAME],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        with st.expander("👀 Превью отчёта «Динамика»", expanded=True):
            workbook = load_workbook(
                BytesIO(st.session_state[DYN_EXCEL_BYTES]),
                data_only=True,
            )

            all_sheet_names = workbook.sheetnames
            preview_sheet_names = _filter_preview_sheets(all_sheet_names)
            default_sheet_names = _smart_default_sheets(preview_sheet_names)

            st.caption(
                f"Листов: **{len(all_sheet_names)}** "
                f"(в превью: **{len(preview_sheet_names)}**). "
                f"Тип: **{report_kind}**, "
                f"ОГХ: **{'Да' if split_by_ogx else 'Нет'}**, "
                f"ИТП: **{'Да' if split_by_itp else 'Нет'}**"
            )

            col_sheets, col_rows, col_cols, col_headers = st.columns([2.2, 1, 1, 1])

            with col_sheets:
                selected_sheet_names = st.multiselect(
                    "Листы для просмотра",
                    options=preview_sheet_names,
                    default=default_sheet_names,
                    key="dyn_prev_sheets",
                )

            with col_rows:
                max_rows = st.number_input(
                    "Строк",
                    min_value=20,
                    max_value=700,
                    value=140,
                    step=10,
                    key="dyn_prev_rows",
                )

            with col_cols:
                max_cols = st.number_input(
                    "Колонок",
                    min_value=10,
                    max_value=150,
                    value=45,
                    step=5,
                    key="dyn_prev_cols",
                )

            with col_headers:
                show_first_row_as_header = st.checkbox(
                    "1-я строка как заголовки",
                    value=False,
                    key="dyn_prev_headers",
                )

            if not selected_sheet_names:
                st.info("Выбери хотя бы один лист для превью.")
            else:
                tabs = st.tabs([f"📄 {sheet_name}" for sheet_name in selected_sheet_names])

                for tab, sheet_name in zip(tabs, selected_sheet_names):
                    with tab:
                        worksheet = workbook[sheet_name]
                        st.caption(f"Размер листа: {worksheet.max_row} × {worksheet.max_column}")

                        preview_rows = []
                        for row in worksheet.iter_rows(
                            min_row=1,
                            max_row=min(worksheet.max_row, int(max_rows)),
                            max_col=min(worksheet.max_column, int(max_cols)),
                            values_only=True,
                        ):
                            preview_rows.append(list(row))

                        preview_df = pd.DataFrame(preview_rows)

                        while (
                            len(preview_df) > 0
                            and preview_df.tail(1).isna().all(axis=1).iloc[0]
                        ):
                            preview_df = preview_df.iloc[:-1]

                        if show_first_row_as_header and len(preview_df) >= 2:
                            header_row = preview_df.iloc[0].astype(str).tolist()
                            body_df = preview_df.iloc[1:].copy()
                            body_df.columns = header_row
                            st.dataframe(body_df, width="stretch")
                        else:
                            st.dataframe(preview_df, width="stretch")
    else:
        st.info("Сформируй отчёт — появятся кнопка скачивания и превью.")

    if st.session_state[DYN_LAST_RUN]:
        st.caption(f"Последняя генерация: {st.session_state[DYN_LAST_RUN]}")