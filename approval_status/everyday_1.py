from __future__ import annotations

import os
import re
import math
import tempfile
import datetime as dt
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from decimal import Decimal, ROUND_HALF_EVEN
from typing import Optional, Iterable, List, Tuple, Dict
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.drawing.image import Image as XLImage




def _norm_header(s: str) -> str:
    """Нормализация заголовка для сравнения."""
    s = "" if s is None else str(s)
    s = s.replace("\u00A0", " ")
    s = s.strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s

def _find_best_column_by_keywords(df: pd.DataFrame, must_have: list[str], any_of: list[str]) -> str | None:
    """
    Ищет колонку, где:
      - присутствуют ВСЕ слова из must_have
      - и хотя бы одно слово из any_of (если any_of не пуст)
    Сравнение по нормализованному заголовку.
    """
    cols = list(df.columns)
    norm_map = {c: _norm_header(c) for c in cols}

    best = None
    best_score = -1

    for c, n in norm_map.items():
        ok_all = all(word in n for word in must_have)
        if not ok_all:
            continue

        ok_any = True
        if any_of:
            ok_any = any(word in n for word in any_of)

        if not ok_any:
            continue

        # скоринг
        score = sum(1 for w in must_have if w in n) + sum(1 for w in any_of if w in n)
        if score > best_score:
            best = c
            best_score = score

    return best

def resolve_ha_column(df: pd.DataFrame, canonical_name: str) -> pd.DataFrame:
    """
    Находит колонку площади (га), если имя плавает, и переименовывает в canonical_name.
    Если canonical_name уже есть — ничего не делает.
    """
    if canonical_name in df.columns:
        return df

    must_have = ["га"]
    any_of = ["сумма", "объем", "объём", "площад", "площадь", "заказ"]

    found = _find_best_column_by_keywords(df, must_have=must_have, any_of=any_of)

    # если не нашли по любому из any_of
    if found is None:
        found = _find_best_column_by_keywords(df, must_have=["га"], any_of=["объем", "объём", "площад", "площадь"])

    if found is None:
        return df

    return df.rename(columns={found: canonical_name})





# Конфиг: имена колонок в БД (ДОЛЖНЫ совпадать с вашей БД)

COL_CONTRACT = "№ Ген. договора"
COL_ORDER = "№ Заказа МГГТ"
COL_STATE = "Состояние (действующий / приостановлен / аннулирован)"
COL_OGH = "Тип объекта ОГХ"
COL_STATUS = "Статус загрузки"
COL_ASU_LOAD = "Дата загрузки в АСУ ОДС"
COL_ASU_APPROVE = "Дата утверждения в АСУ ОДС (МГГТ)"
COL_HA = "Сумма Объем заказа, га"
COL_STAGE = "Номер этапа МГГТ"
COL_KIND = "Актуализация / Первичное обследование"
COL_OIV = "ОИВ"
COL_BALANCE = "Балансодержатель"
COL_OBJ_NAME = "Наименование объекта"
COL_ASU_LOAD_ASD = "Дата загрузки в АСУ ОДС (АСД)"

EXPORT_COLS = [
    COL_ORDER,
    COL_STAGE,
    COL_KIND,
    COL_OIV,
    COL_BALANCE,
    COL_OGH,
    COL_OBJ_NAME,
    COL_ASU_LOAD_ASD,
    COL_ASU_LOAD,
    COL_HA,
    COL_STATUS,
]

REQUIRED_COLS = [
    # используется в логике фильтрации/масок/агрегаций
    COL_CONTRACT, COL_ORDER, COL_STATE, COL_OGH, COL_STATUS,
    COL_ASU_LOAD, COL_ASU_APPROVE, COL_HA,
    COL_STAGE, COL_KIND, COL_OIV, COL_BALANCE, COL_OBJ_NAME, COL_ASU_LOAD_ASD,
]

OGH_ORDER = ["ДТ", "ОО", "ОДХ"]


REPORT_WIDTH_COLS = 10

# Жирные вертикальные разделители после:
#  - Тип ОГХ (B) -> offset 0
#  - Всего (C)   -> offset 1
#  - % утв (E)   -> offset 3
#  - % откл (G)  -> offset 5
#  - % рассм (I) -> offset 7
THICK_VLINE_AFTER_OFFSETS = [0, 1, 3, 5, 7]

REJECTED_STATUSES = {
    "Запрос обрабатывается",
    "Получен ответ об ошибке",
    "Проект был отклонен",
    "Проект утвержден",
    "Задача отправлена в АСУ ОДС",
    "Ошибка обработки в АСУ ОДС",
}
REVIEW_STATUSES = {
    "Акт подписан",
    "Отправлен на согласование",
    "Получен ответ",
    "Согласован с внешней системой",
    "Объект создан в АСУ ОДС",
}

# Цвета колонок
FILL_APPROVED = PatternFill("solid", fgColor="DDF6D6")  # утверждено + %
FILL_REJECTED = PatternFill("solid", fgColor="FFE699")  # отклонено + %
FILL_REVIEW = PatternFill("solid", fgColor="F8CBAD")  # на рассмотрении + %

# Утилиты
def round_gauss_2(x: float) -> float:
    return float(Decimal(str(x)).quantize(Decimal("0.01"), rounding=ROUND_HALF_EVEN))


def contract_suffix_year(val) -> Optional[int]:
    if pd.isna(val):
        return None
    s = str(val).strip()

    m = re.search(r"-(\d{2})\s*\D*$", s)
    if m:
        return int(m.group(1))

    m = re.search(r"(\d{2})\s*\D*$", s)
    if m:
        return int(m.group(1))

    return None


def normalize_cols(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df


def ensure_columns(df: pd.DataFrame, required: Iterable[str]) -> None:
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise KeyError(f"Не найдены обязательные колонки в БД: {missing}")


def detect_header_row(excel_path: str, required_cols: List[str], scan_rows: int = 80) -> int:
    raw = pd.read_excel(excel_path, header=None, nrows=scan_rows)
    required_set = set(required_cols)

    best_row = 0
    best_hits = -1
    for r in range(raw.shape[0]):
        row_vals = [str(x).strip() for x in raw.iloc[r].tolist() if pd.notna(x)]
        hits = sum(1 for x in row_vals if x in required_set)
        if hits > best_hits:
            best_hits = hits
            best_row = r
        if hits >= max(3, int(len(required_cols) * 0.6)):
            return r

    return best_row


def load_db_excel(excel_path: str) -> pd.DataFrame:
    hdr = detect_header_row(excel_path, REQUIRED_COLS, scan_rows=80)
    df = pd.read_excel(excel_path, header=hdr)
    df = normalize_cols(df)

    # авто-поиск колонки "Сумма Объем заказа, га"
    df = resolve_ha_column(df, COL_HA)

    ensure_columns(df, REQUIRED_COLS)
    return df

def available_year_suffixes(df: pd.DataFrame) -> List[int]:
    years = (
        df[COL_CONTRACT]
        .apply(contract_suffix_year)
        .dropna()
        .astype(int)
        .unique()
        .tolist()
    )
    return sorted(set(years))


def fill_dosyem_status_from_parent(dfp: pd.DataFrame) -> pd.DataFrame:
    """
    Для заказов-досъёмов (№ Заказа МГГТ заканчивается на 'Д'):
    если Статус загрузки пустой -> подставляем Статус загрузки родителя.

    Родитель определяется по совпадению префикса первых 3 сегментов до '/'.
    """
    if COL_ORDER not in dfp.columns or COL_STATUS not in dfp.columns:
        return dfp

    x = dfp.copy()

    order = x[COL_ORDER].astype(str).str.strip()
    is_dos = order.str.upper().str.endswith("Д")

    def prefix3(s: str) -> str:
        parts = str(s).split("/")
        return "/".join(parts[:3]) if len(parts) >= 3 else str(s)

    p3 = order.map(prefix3)

    # родители = всё, что НЕ досъём
    parents = x.loc[~is_dos, [COL_STATUS]].copy()
    parents["_p3"] = p3.loc[~is_dos].values

    # берём статус родителя (приоритет: непустой/не NaN)
    parents["_st"] = parents[COL_STATUS]
    parents_ok = parents.loc[parents["_st"].notna()].copy()

    parent_status_map = parents_ok.drop_duplicates("_p3")[["_p3", "_st"]].set_index("_p3")["_st"]

    # у досъёмов заполняем только если пусто
    dos_need = is_dos & x[COL_STATUS].isna()
    if dos_need.any():
        x.loc[dos_need, COL_STATUS] = p3.loc[dos_need].map(parent_status_map)

    return x

def prepare_filtered_df(df: pd.DataFrame, year_suffix: int) -> Tuple[pd.DataFrame, Dict[str, int]]:
    stats: Dict[str, int] = {}

    dfp = df.copy()
    stats["0_всего_строк"] = len(dfp)

    suffix = dfp[COL_CONTRACT].apply(contract_suffix_year)
    dfp = dfp.loc[suffix == year_suffix].copy()
    stats["1_после_года"] = len(dfp)

    bad_letters = (
        dfp[COL_ORDER]
        .astype(str)
        .str.upper()
        .str.contains(r"[РКВRKVЮ]", regex=True, na=False)
    )
    dfp = dfp.loc[~bad_letters].copy()
    stats["2_после_рквю"] = len(dfp)

    #  действующий + приостановленные
    state_norm = (
        dfp[COL_STATE]
        .astype(str)
        .str.strip()
        .str.lower()
    )
    dfp = dfp.loc[state_norm.isin(["действующий", "приостановленные"])].copy()
    stats["3_после_состояния"] = len(dfp)

    ogh_norm = (
        dfp[COL_OGH]
        .astype(str)
        .str.strip()
        .str.upper()
    )
    dfp = dfp.loc[ogh_norm.isin(OGH_ORDER)].copy()
    stats["4_после_огх"] = len(dfp)

    dfp["_ОГХ_НОРМ"] = dfp[COL_OGH].astype(str).str.strip().str.upper()

    # ДОСЪЁМЫ: если статус пустой берём статус родителя
    dfp = fill_dosyem_status_from_parent(dfp)

    return dfp, stats

def build_masks(dfp: pd.DataFrame):
    approved_mask = dfp[COL_ASU_LOAD_ASD].notna() & dfp[COL_ASU_APPROVE].notna()

    status = dfp[COL_STATUS]
    empty_status_mask = status.isna()

    rejected_mask = (~approved_mask) & (empty_status_mask | status.isin(REJECTED_STATUSES))
    review_mask = (~approved_mask) & status.isin(REVIEW_STATUSES)
    return approved_mask, rejected_mask, review_mask


def aggregate_counts(dfp: pd.DataFrame, approved_mask, rejected_mask, review_mask):
    grp_col = "_ОГХ_НОРМ" if "_ОГХ_НОРМ" in dfp.columns else COL_OGH

    total = dfp.groupby(grp_col).size().reindex(OGH_ORDER, fill_value=0)
    appr = dfp.loc[approved_mask].groupby(grp_col).size().reindex(OGH_ORDER, fill_value=0)
    rej = dfp.loc[rejected_mask].groupby(grp_col).size().reindex(OGH_ORDER, fill_value=0)
    rev = dfp.loc[review_mask].groupby(grp_col).size().reindex(OGH_ORDER, fill_value=0)
    remain = total - appr
    return total, appr, rej, rev, remain


def aggregate_ha(dfp: pd.DataFrame, approved_mask, rejected_mask, review_mask):
    grp_col = "_ОГХ_НОРМ" if "_ОГХ_НОРМ" in dfp.columns else COL_OGH

    ha = pd.to_numeric(dfp[COL_HA], errors="coerce").fillna(0.0)
    df2 = dfp.copy()
    df2["_ha"] = ha

    total = df2.groupby(grp_col)["_ha"].sum().reindex(OGH_ORDER, fill_value=0.0)
    appr = df2.loc[approved_mask].groupby(grp_col)["_ha"].sum().reindex(OGH_ORDER, fill_value=0.0)
    rej = df2.loc[rejected_mask].groupby(grp_col)["_ha"].sum().reindex(OGH_ORDER, fill_value=0.0)
    rev = df2.loc[review_mask].groupby(grp_col)["_ha"].sum().reindex(OGH_ORDER, fill_value=0.0)
    remain = total - appr
    return total, appr, rej, rev, remain


def build_unique_path(base_dir: str, year_suffix: int, report_date: dt.date) -> str:
    date_str = report_date.strftime("%d.%m.%Y")
    base = os.path.join(
        base_dir,
        f"Статус утверждения гз {year_suffix} {date_str}"
    )

    for i in range(0, 200):
        suf = "" if i == 0 else f" ({i})"
        p = f"{base}{suf}.xlsx"
        if not os.path.exists(p):
            return p

    raise RuntimeError("Не удалось подобрать уникальное имя файла для отчёта.")



#  Border helpers

def _set_border_sides(cell, *, left=None, right=None, top=None, bottom=None):
    b = cell.border
    cell.border = Border(
        left=left if left is not None else b.left,
        right=right if right is not None else b.right,
        top=top if top is not None else b.top,
        bottom=bottom if bottom is not None else b.bottom,
        diagonal=b.diagonal,
        diagonal_direction=b.diagonal_direction,
        outline=b.outline,
        vertical=b.vertical,
        horizontal=b.horizontal,
    )


def draw_outer_border(ws, top_row: int, left_col: int, bottom_row: int, right_col: int) -> None:
    thick = Side(style="medium")

    for c in range(left_col, right_col + 1):
        _set_border_sides(ws.cell(top_row, c), top=thick)
        _set_border_sides(ws.cell(bottom_row, c), bottom=thick)

    for r in range(top_row, bottom_row + 1):
        _set_border_sides(ws.cell(r, left_col), left=thick)
        _set_border_sides(ws.cell(r, right_col), right=thick)


def draw_thick_bottom_row(ws, row: int, left_col: int, right_col: int) -> None:
    thick = Side(style="medium")
    for c in range(left_col, right_col + 1):
        _set_border_sides(ws.cell(row, c), bottom=thick)


def draw_thick_vertical_after(ws, top_row: int, bottom_row: int, col_idx: int) -> None:
    thick = Side(style="medium")
    for r in range(top_row, bottom_row + 1):
        _set_border_sides(ws.cell(r, col_idx), right=thick)


def fill_columns_block(ws, row_top: int, row_bottom: int, col_indices: List[int], fill: PatternFill) -> None:
    """Красим диапазон по списку абсолютных индексов колонок."""
    for r in range(row_top, row_bottom + 1):
        for c in col_indices:
            ws.cell(r, c).fill = fill



#  Выгрузка df -> лист (для 4 листов статусов)

def _to_excel_value(v):
    """Аккуратно превращаем pandas-значения в то, что нормально пишет openpyxl."""
    if pd.isna(v):
        return None
    if isinstance(v, (pd.Timestamp, dt.datetime)):
        return v.to_pydatetime() if hasattr(v, "to_pydatetime") else v
    if isinstance(v, dt.date):
        return v
    return v


def write_df_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame, cols: List[str]) -> None:
    ws = wb.create_sheet(title=sheet_name)

    # Берём только нужные колонки
    dfx = df.loc[:, cols].copy()

    # Заголовки
    header_font = Font(bold=True)
    for j, col in enumerate(cols, start=1):
        cell = ws.cell(1, j, col)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Данные: пишем по позициям
    values = dfx.to_numpy()
    for i in range(values.shape[0]):
        for j in range(values.shape[1]):
            ws.cell(i + 2, j + 1).value = _to_excel_value(values[i, j])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Автоширина
    for j, col in enumerate(cols, start=1):
        max_len = len(str(col))
        sample_n = min(len(dfx), 5000)
        for k in range(sample_n):
            vv = dfx.iloc[k, j - 1]
            if pd.isna(vv):
                continue
            max_len = max(max_len, len(str(vv)))
        ws.column_dimensions[ws.cell(1, j).column_letter].width = min(max(10, max_len + 2), 60)


def make_status_pie_png(
    pie_png: str,
    sizes: list,
    labels: list,
    colors: list,
    title: str = "",
    small_share_threshold: float = 0.05,
):
    """
    Рисует круговую диаграмму:
      - сектора >= small_share_threshold подписываются выносными линиями вокруг круга
      - сектора < small_share_threshold НЕ подписываются выносом, а выводятся списком справа:
            "Название 5.0%"

    sizes: значения по секторам
    labels: названия секторов (строки)
    colors: цвета секторов (список цветов в формате matplotlib)
    pie_png: путь сохранения PNG
    """

    # нормализация входа
    sizes = [0 if v is None else float(v) for v in sizes]
    total = sum(sizes)

    # если всё нулевое — рисуем заглушку
    if total <= 0:
        fig, ax = plt.subplots(figsize=(8.6, 4.6), dpi=160)
        ax.axis("off")
        ax.text(0.5, 0.5, "Нет данных для построения диаграммы",
                ha="center", va="center", fontsize=14, fontweight="bold")
        if title:
            ax.set_title(title, fontsize=14, fontweight="bold", pad=14)
        fig.savefig(pie_png, bbox_inches="tight", pad_inches=0.6)
        plt.close(fig)
        return

    shares = [v / total for v in sizes]

    #  делаем справа место под список маленьких долей
    fig, ax = plt.subplots(figsize=(9.6, 5.0), dpi=160)
    fig.subplots_adjust(left=0.06, right=0.72, top=0.88, bottom=0.08)

    wedges, texts, autotexts = ax.pie(
        sizes,
        labels=None,
        colors=colors,
        startangle=90,
        counterclock=False,
        autopct=lambda p: f"{p:.1f}%",
        pctdistance=0.72,
        wedgeprops={"linewidth": 1.2, "edgecolor": "white"},
    )

    ax.axis("equal")

    if title:
        ax.set_title(title, fontsize=14, fontweight="bold", pad=14)

    # скрываем стандартные проценты внутри
    for t in autotexts:
        t.set_visible(False)

    #  разнесение: большие доли -> вынос, маленькие -> список справа
    small_items = []

    for i, w in enumerate(wedges):
        pct_txt = f"{shares[i] * 100:.1f}%"
        label_txt = str(labels[i]).strip()

        if shares[i] < small_share_threshold:
            small_items.append(f"{label_txt} {pct_txt}")
            continue

        # угол сектора
        ang = (w.theta2 + w.theta1) / 2.0
        x = math.cos(math.radians(ang))
        y = math.sin(math.radians(ang))

        ha = "left" if x >= 0 else "right"
        x_text = 1.55 if x >= 0 else -1.55
        y_text = 1.10 * y

        ax.annotate(
            f"{label_txt}\n{pct_txt}",
            xy=(0.92 * x, 0.92 * y),
            xytext=(x_text, y_text),
            ha=ha,
            va="center",
            fontsize=12,
            fontweight="bold",
            arrowprops={"arrowstyle": "-", "lw": 1.2},
        )


    if small_items:
        small_title = "Малые доли:"
        block = small_title + "\n" + "\n".join(small_items)

        fig.text(
            0.76, 0.50, block,
            ha="left", va="center",
            fontsize=12,
            fontweight="bold"
        )

    fig.savefig(pie_png, bbox_inches="tight", pad_inches=0.6)
    plt.close(fig)


def insert_status_pie_to_ws(
    ws,
    anchor_row: int,
    anchor_col: int,
    year_suffix: int,
    approved: float,
    rejected: float,
    review: float,
) -> tuple[bool, str | None]:

    """
    Строит PNG pie по approved/rejected/review и вставляет картинку на ws в anchor.
    Возвращает True
    """
    pie_png = os.path.join(
        tempfile.gettempdir(),
        f"status_pie_gz{year_suffix}_{int(dt.datetime.now().timestamp())}.png"
    )

    total = float(approved) + float(rejected) + float(review)
    if total <= 0:
        return False, None

    sizes = [approved, review, rejected]
    labels = [
        "УТВЕРЖДЕНО",
        "НА РАССМОТРЕНИИ\nу балансодержателя",
        "ОТКЛОНЕНО\n(Направлено на\nисправление в МГГТ)",
    ]
    explode = (0.03, 0.03, 0.05)

    fig, ax = plt.subplots(figsize=(10, 6), dpi=150)

    colors = [
        _fill_to_hex(FILL_APPROVED),
        _fill_to_hex(FILL_REVIEW),
        _fill_to_hex(FILL_REJECTED),
    ]

    wedges, texts, autotexts = ax.pie(
        sizes,
        labels=None,
        colors=colors,
        autopct=lambda p: f"{p:.1f}%",
        startangle=20,
        explode=explode,
        pctdistance=0.62,
        wedgeprops={"edgecolor": "black", "linewidth": 1.2},
    )
    ax.axis("equal")

    # подписи
    total = sum(sizes)
    shares = [v / total if total else 0 for v in sizes]

    #  если один сектор = 100%
    nonzero_cnt = sum(1 for s in shares if s > 0.00001)
    if nonzero_cnt <= 1 or max(shares) >= 0.999:
        ax.clear()

        # рисуем круг без подписей
        wedges, texts = ax.pie(
            sizes,
            labels=None,
            colors=colors,
            autopct=None,  # важно: только 2 значения
            startangle=20,
            explode=[0] * len(sizes),
            wedgeprops={"edgecolor": "black", "linewidth": 1.2},
        )
        ax.axis("equal")

        # подпись в центре
        idx = int(np.argmax(sizes))
        label_txt = labels[idx] if labels else ""
        ax.text(
            0, 0,
            f"{label_txt}\n100%",
            ha="center",
            va="center",
            fontsize=14,
            fontweight="bold",
        )

        fig.savefig(pie_png, bbox_inches="tight", pad_inches=0.4)
        plt.close(fig)

        if (not os.path.exists(pie_png)) or os.path.getsize(pie_png) < 1000:
            return False, None

        img = XLImage(pie_png)
        anchor_cell = f"{ws.cell(1, anchor_col).column_letter}{anchor_row}"
        ws.add_image(img, anchor_cell)
        return True, pie_png

    small_threshold = 0.05
    small_items = []

    for i, w in enumerate(wedges):
        if shares[i] <= 0:
            continue

        pct_txt = f"{shares[i] * 100:.1f}%"
        label_txt = str(labels[i]).strip()

        # маленькие доли -> список справа
        if shares[i] < small_threshold:
            small_items.append(f"{label_txt} {pct_txt}")
            continue

        # большие доли -> выносная подпись с линией
        ang = (w.theta2 + w.theta1) / 2.0
        x = math.cos(math.radians(ang))
        y = math.sin(math.radians(ang))

        ha = "left" if x >= 0 else "right"
        x_text = 1.55 if x >= 0 else -1.55
        y_text = 1.10 * y

        ax.annotate(
            f"{label_txt}\n{pct_txt}",
            xy=(0.92 * x, 0.92 * y),
            xytext=(x_text, y_text),
            ha=ha,
            va="center",
            fontsize=12,
            fontweight="bold",
            arrowprops={"arrowstyle": "-", "lw": 1.2},
        )

    # маленькие доли: текстовый список справа от круга
    if small_items:
        fig.subplots_adjust(left=0.24, right=0.62, top=0.95, bottom=0.07)

        block = "Малые доли:\n" + "\n".join(small_items)
        fig.text(
            0.80, 0.50, block,
            ha="left", va="center",
            fontsize=12,
            fontweight="bold"
        )

    for t in autotexts:
        t.set_visible(False)


    plt.tight_layout()
    fig.savefig(pie_png, bbox_inches="tight", pad_inches=0.4)
    plt.close(fig)

    # если png не появился — выходим
    if (not os.path.exists(pie_png)) or os.path.getsize(pie_png) < 1000:
        return False, None

    # вставляем в Excel
    img = XLImage(pie_png)
    anchor_cell = f"{ws.cell(1, anchor_col).column_letter}{anchor_row}"
    ws.add_image(img, anchor_cell)

    return True, pie_png


# 4) Excel-таблица (лист "Отчет")


def write_table(
    ws,
    start_row: int,
    start_col: int,
    title_year_suffix: int,
    report_date: dt.date,
    unit_label: str,
    total, appr, rej, rev, remain,
    is_float: bool,
    round_ga: bool,
) -> int:
    thin = Side(style="thin")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    W = REPORT_WIDTH_COLS  # логическая ширина
    COL_SPAN = 2
    left_col = start_col
    right_col = start_col + W * COL_SPAN - 1

    def col_of(off: int) -> int:
        """Левая физическая колонка для логической колонки """
        return left_col + off * COL_SPAN

    def merge_row_cols(r: int):
        """merge двух физ.колонок в одну для каждой логической колонки в строке r."""
        for off in range(W):
            c0 = col_of(off)
            ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c0 + 1)

    bold = Font(bold=True)
    normal = Font(bold=False)

    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def set_thin_row(r_abs: int):
        for c in range(left_col, right_col + 1):
            ws.cell(r_abs, c).border = thin_border

    def write_abs(cell_obj, v, *, bold_numbers: bool):
        if is_float:
            vv = float(v)
            if round_ga:
                vv = round_gauss_2(vv)
                cell_obj.number_format = "0.0"
            cell_obj.value = vv
        else:
            cell_obj.value = int(v)
        cell_obj.font = bold if bold_numbers else normal

    def write_percent(cell_obj, num, den, *, bold_numbers: bool):
        den_f = float(den) if den is not None else 0.0
        if den_f == 0.0:
            cell_obj.value = None
            cell_obj.font = bold if bold_numbers else normal
            return
        val = float(num) / den_f
        cell_obj.value = val
        cell_obj.number_format = "0.0%" if abs(val) >= 0.01 else "0.0%"
        cell_obj.font = bold if bold_numbers else normal

    # Row 1
    ws.merge_cells(start_row=start_row, start_column=left_col, end_row=start_row, end_column=right_col)
    c0 = ws.cell(start_row, left_col)
    unit_suffix = unit_label.lower()
    c0.value = (
        f"Статус утверждения ГЗ-{title_year_suffix}"
        f"от {report_date:%d.%m.%Y}, {unit_suffix}"
    )
    c0.alignment = center
    c0.font = bold
    set_thin_row(start_row)
    draw_thick_bottom_row(ws, start_row, left_col, right_col)
    ws.row_dimensions[start_row].height = 30

    # Row 2
    ws.merge_cells(start_row=start_row + 1, start_column=left_col, end_row=start_row + 1, end_column=right_col)
    c1 = ws.cell(start_row + 1, left_col)
    c1.value = "Статус утверждения в АСУ ОДС"
    c1.alignment = center
    c1.font = bold
    set_thin_row(start_row + 1)
    draw_thick_bottom_row(ws, start_row + 1, left_col, right_col)
    ws.row_dimensions[start_row + 1].height = 22


    hdr_r = start_row + 2

    merge_row_cols(hdr_r)

    ch = ws.cell(hdr_r, col_of(0))
    ch.value = "Тип объекта ОГХ"
    ch.alignment = center
    ch.font = bold

    headers = [
        (f"Всего ГЗ {title_year_suffix}", 1),
        ("УТВЕРЖДЕНО", 2),
        ("% утверждено", 3),
        ("ОТКЛОНЕНО\n(Направлено на исправление в МГГТ)", 4),
        ("% отклонения", 5),
        ("НА РАССМОТРЕНИИ\nу балансодержателя", 6),
        ("% на\nрассмотрении", 7),
        ("ОСТАЛОСЬ\nУТВЕРДИТЬ", 8),
        ("%\nнеутвержденных", 9),
    ]
    for text, off in headers:
        cell = ws.cell(hdr_r, col_of(off))
        cell.value = text
        cell.alignment = center
        cell.font = bold

    set_thin_row(hdr_r)
    draw_thick_bottom_row(ws, hdr_r, left_col, right_col)
    ws.row_dimensions[hdr_r].height = 80

    data_start_r = start_row + 3

    for i, t in enumerate(OGH_ORDER):
        r = data_start_r + i
        merge_row_cols(r)

        ct = ws.cell(r, left_col)
        ct.value = t
        ct.alignment = center
        ct.font = normal

        T = total[t]
        A = appr[t]
        Rj = rej[t]
        V = rev[t]
        L = remain[t]

        abs_map = {
            col_of(1): T,
            col_of(2): A,
            col_of(4): Rj,
            col_of(6): V,
            col_of(8): L,
        }
        for c_abs, v in abs_map.items():
            cobj = ws.cell(r, c_abs)
            write_abs(cobj, v, bold_numbers=False)
            cobj.alignment = center

        pct_map = {
            col_of(3): (A, T),
            col_of(5): (Rj, T),
            col_of(7): (V, T),
            col_of(9): (L, T),
        }
        for c_abs, (num, den) in pct_map.items():
            cobj = ws.cell(r, c_abs)
            write_percent(cobj, num, den, bold_numbers=False)
            cobj.alignment = center

        for c_abs in range(left_col, right_col + 1):
            cell = ws.cell(r, c_abs)
            cell.border = thin_border
            if cell.alignment is None:
                cell.alignment = center

        ws.row_dimensions[r].height = 22

    # Σ row
    sum_r = data_start_r + 3
    merge_row_cols(sum_r)

    cs = ws.cell(sum_r, left_col)
    cs.value = "ИТОГО"
    cs.alignment = Alignment(horizontal="center", vertical="center")
    cs.font = bold

    T_sum = total.sum()
    A_sum = appr.sum()
    R_sum = rej.sum()
    V_sum = rev.sum()
    L_sum = remain.sum()

    abs_map = {
        col_of(1): total.sum(),  # Всего
        col_of(2): appr.sum(),  # Утверждено
        col_of(4): rej.sum(),  # Отклонено
        col_of(6): rev.sum(),  # На рассмотрении
        col_of(8): remain.sum(),  # Осталось
    }
    for c_abs, v in abs_map.items():
        cobj = ws.cell(sum_r, c_abs)
        write_abs(cobj, v, bold_numbers=True)
        cobj.alignment = center

    pct_map = {
        col_of(3): (appr.sum(), total.sum()),
        col_of(5): (rej.sum(), total.sum()),
        col_of(7): (rev.sum(), total.sum()),
        col_of(9): (remain.sum(), total.sum()),
    }
    for c_abs, (num, den) in pct_map.items():
        cobj = ws.cell(sum_r, c_abs)
        write_percent(cobj, num, den, bold_numbers=True)
        cobj.alignment = center

    for c_abs in range(left_col, right_col + 1):
        cell = ws.cell(sum_r, c_abs)
        cell.border = thin_border
        if cell.alignment is None:
            cell.alignment = center
        if cell.value is None:
            cell.font = bold

    ws.row_dimensions[sum_r].height = 24

    # Закраска колонок
    fill_columns_block(
        ws, row_top=hdr_r, row_bottom=sum_r,
        col_indices=[col_of(2), col_of(2) + 1, col_of(3), col_of(3) + 1],
        fill=FILL_APPROVED
    )
    fill_columns_block(
        ws, row_top=hdr_r, row_bottom=sum_r,
        col_indices=[col_of(4), col_of(4) + 1, col_of(5), col_of(5) + 1],
        fill=FILL_REJECTED
    )
    fill_columns_block(
        ws, row_top=hdr_r, row_bottom=sum_r,
        col_indices=[col_of(6), col_of(6) + 1, col_of(7), col_of(7) + 1],
        fill=FILL_REVIEW
    )

    # Вертикальные жирные линии
    top_r = start_row
    bottom_r = sum_r
    for off in THICK_VLINE_AFTER_OFFSETS:
        draw_thick_vertical_after(ws, top_r, bottom_r, col_of(off) + 1)

    return sum_r

def _fill_to_hex(fill: PatternFill) -> str:

    rgb = getattr(getattr(fill, "fgColor", None), "rgb", None)
    if not rgb:
        return "#CCCCCC"
    rgb = str(rgb)
    if len(rgb) == 8:
        rgb = rgb[2:]
    if len(rgb) != 6:
        return "#CCCCCC"
    return f"#{rgb}"


def build_report_excel(df: pd.DataFrame, year_suffix: int, out_path: str, report_date: Optional[dt.date] = None) -> Tuple[str, str]:
    if report_date is None:
        report_date = dt.date.today()

    df = normalize_cols(df)
    ensure_columns(df, REQUIRED_COLS)

    dfp, stats = prepare_filtered_df(df, year_suffix)
    approved_mask, rejected_mask, review_mask = build_masks(dfp)
    remain_mask = ~approved_mask



    #  НОВАЯ ДИАГНОСТИКА:
    # Дата загрузки (АСД) пустая, а дата утверждения (МГГТ) НЕ пустая

    bad_dates_mask = dfp[COL_ASU_LOAD_ASD].isna() & dfp[COL_ASU_APPROVE].notna()
    bad_orders = (
        dfp.loc[bad_dates_mask, COL_ORDER]
        .dropna()
        .astype(str)
        .str.strip()
        .unique()
        .tolist()
    )
    bad_orders.sort()

    max_show = 200
    if len(bad_orders) == 0:
        bad_orders_block = "Нет.\n"
    else:
        shown = bad_orders[:max_show]
        rest = len(bad_orders) - len(shown)
        bad_orders_block = ", ".join(shown)
        if rest > 0:
            bad_orders_block += f", ... и еще {rest}"

    debug_text = (
        f"Всего строк в БД: {stats.get('0_всего_строк', len(df))}\n"
        f"После фильтра ГОД: {stats.get('1_после_года', 0)}\n"
        f"После фильтра РКВ: {stats.get('2_после_ркв', 0)}\n"
        f"После фильтра СОСТОЯНИЕ: {stats.get('3_после_состояния', 0)}\n"
        f"После фильтра ОГХ: {stats.get('4_после_огх', 0)}\n"
        f"Утверждено (по датам): {int(approved_mask.sum())}\n"
        f"Отклонено (по статусам МГГТ, только НЕутверждённые): {int(rejected_mask.sum())}\n"
        f"На рассмотрении (по статусам БД, только НЕутверждённые): {int(review_mask.sum())}\n"
        f"Осталось утвердить (все НЕутверждённые): {int(remain_mask.sum())}\n"
        f"\n"
        f"Проблемные записи (утверждение МГГТ есть, а загрузки АСД нет): {len(bad_orders)}\n"
        f"Список № Заказа МГГТ: {bad_orders_block}\n"
    )

    total_c, appr_c, rej_c, rev_c, rem_c = aggregate_counts(dfp, approved_mask, rejected_mask, review_mask)
    total_h, appr_h, rej_h, rev_h, rem_h = aggregate_ha(dfp, approved_mask, rejected_mask, review_mask)

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет"

    # Ширины под сетку (A..K)
    widths = {
        "A": 4,
        "B": 16,
        "C": 13.00,
        "D": 25.00, "E": 25.00, "F": 25.00, "G": 25.00,
        "H": 25.00, "I": 25.00, "J": 25.00, "K": 25.00,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    start_row_1 = 2
    start_col_1 = 2
    last_row_1 = write_table(
        ws, start_row_1, start_col_1,
        year_suffix, report_date,
        "ШТ",
        total_c, appr_c, rej_c, rev_c, rem_c,
        is_float=False, round_ga=False
    )
    draw_outer_border(ws, start_row_1, start_col_1, last_row_1, start_col_1 + REPORT_WIDTH_COLS - 1)

    start_row_2 = last_row_1 + 4
    start_col_2 = start_col_1
    last_row_2 = write_table(
        ws, start_row_2, start_col_2,
        year_suffix, report_date,
        "ГА",
        total_h, appr_h, rej_h, rev_h, rem_h,
        is_float=True, round_ga=True
    )
    draw_outer_border(ws, start_row_2, start_col_2, last_row_2, start_col_2 + REPORT_WIDTH_COLS - 1)


    #  Pie-график по ШТ

    pie_png = os.path.join(
        tempfile.gettempdir(),
        f"status_pie_gz{year_suffix}_{int(dt.datetime.now().timestamp())}.png"
    )

    ok_pie, pie_png = insert_status_pie_to_ws(
        ws,
        anchor_row=last_row_2 + 2,
        anchor_col=start_col_1,
        year_suffix=year_suffix,
        approved=float(appr_c.sum()),
        rejected=float(rej_c.sum()),
        review=float(rev_c.sum()),
    )

    # 4 листа детализации по тем же маскам, что и таблица

    df_ok = dfp.loc[approved_mask, EXPORT_COLS].copy()
    df_rej = dfp.loc[rejected_mask, EXPORT_COLS].copy()
    df_rev = dfp.loc[review_mask, EXPORT_COLS].copy()
    df_rem = dfp.loc[remain_mask, EXPORT_COLS].copy()

    sort_cols = [COL_OGH, COL_ORDER, COL_STAGE]
    for dfx in (df_ok, df_rej, df_rev, df_rem):
        existing = [c for c in sort_cols if c in dfx.columns]
        if existing:
            dfx.sort_values(existing, inplace=True, kind="mergesort")

    write_df_sheet(wb, "Утверждено", df_ok, EXPORT_COLS)
    write_df_sheet(wb, "Отклонено", df_rej, EXPORT_COLS)
    write_df_sheet(wb, "На рассмотрении", df_rev, EXPORT_COLS)
    write_df_sheet(wb, "Осталось утвердить", df_rem, EXPORT_COLS)

    wb.save(out_path)

    return out_path, debug_text


# Закрытие сводок

COL_EXECUTOR = "Исполнитель загрузки"

def _order_prefix3(order_val: str) -> str:
    """Ключ родителя/досъема: первые 3 сегмента до 3-го слэша."""
    s = str(order_val).strip()
    parts = s.split("/")
    return "/".join(parts[:3])

def _is_dosyem(order_val: str) -> bool:
    """Досъем: № Заказа МГГТ заканчивается на /Д или просто на Д."""
    s = str(order_val).strip().upper()
    return s.endswith("/Д") or s.endswith("Д")

def build_closure_issues(dfp: pd.DataFrame) -> pd.DataFrame:
    """
    Возвращает DataFrame проблемных заказов для 'закрытия сводок'.

    Группа 1 (досъемы):
      - заказ-досъем (оканчивается на Д) НЕ утвержден (нет даты в COL_ASU_APPROVE)
      - у родителя (по префиксу 3 сегмента) ЕСТЬ утверждение:
            либо дата (COL_ASU_APPROVE),
            либо COL_STATUS == 'Проект утвержден' (если даты нет)
      - В итоговой строке досъема:
            Исполнитель загрузки = от родителя
            Статус загрузки      = от родителя
            Дата утверждения МГГТ = от досъема (пустая — это нормально)

    Группа 2 (ошибка дат):
      - дата утверждения МГГТ есть
      - дата загрузки (АСД) пустая
    """
    required = [COL_ORDER, COL_STATUS, COL_ASU_LOAD_ASD, COL_ASU_APPROVE]
    for c in required:
        if c not in dfp.columns:
            return pd.DataFrame()

    x = dfp.copy()

    # Группа 2: утверждено, но нет загрузки АСД
    mask_bad_dates = x[COL_ASU_APPROVE].notna() & x[COL_ASU_LOAD_ASD].isna()
    bad_dates = x.loc[mask_bad_dates].copy()
    if len(bad_dates) > 0:
        bad_dates_out = pd.DataFrame({
            "Тип проблемы": "Утверждено без загрузки АСД",
            COL_ORDER: bad_dates[COL_ORDER],
            "№ Родительского заказа": pd.NA,
            "Дата утверждения родителя": pd.NA,
            "Статус родителя": pd.NA,
            COL_ASU_APPROVE: bad_dates[COL_ASU_APPROVE],
            COL_ASU_LOAD_ASD: bad_dates[COL_ASU_LOAD_ASD],
            COL_EXECUTOR: bad_dates[COL_EXECUTOR] if COL_EXECUTOR in bad_dates.columns else pd.NA,
            COL_STATUS: bad_dates[COL_STATUS],
        })
    else:
        bad_dates_out = pd.DataFrame()

    # Группа 1: досъемы /Д
    order_series = x[COL_ORDER].astype(str).str.strip()
    dos_mask = order_series.apply(_is_dosyem)

    dos = x.loc[dos_mask].copy()
    parents = x.loc[~dos_mask].copy()

    dos_out = pd.DataFrame()
    if len(dos) > 0 and len(parents) > 0:
        dos["_p3"] = dos[COL_ORDER].astype(str).apply(_order_prefix3)
        parents["_p3"] = parents[COL_ORDER].astype(str).apply(_order_prefix3)

        # родитель "утвержден": дата есть ИЛИ статус "Проект утвержден"
        parent_approved_mask = parents[COL_ASU_APPROVE].notna() | (parents[COL_STATUS] == "Проект утвержден")

        # безопасно выбираем колонки родителя (исполнитель может отсутствовать)
        parent_cols = [COL_ORDER, COL_ASU_APPROVE, COL_STATUS, "_p3"]
        if COL_EXECUTOR in parents.columns:
            parent_cols.insert(3, COL_EXECUTOR)  # перед _p3

        parents_ok = parents.loc[parent_approved_mask, parent_cols].copy()

        # "лучший" родитель на префикс:
        # 1) где есть дата утверждения
        # 2) если дат несколько — берём с максимальной датой
        parents_ok["_has_date"] = parents_ok[COL_ASU_APPROVE].notna().astype(int)
        parents_ok.sort_values(["_p3", "_has_date", COL_ASU_APPROVE], ascending=[True, False, False], inplace=True)
        parents_best = parents_ok.drop_duplicates("_p3", keep="first").copy()

        parents_best.rename(columns={
            COL_ORDER: "№ Родительского заказа",
            COL_ASU_APPROVE: "Дата утверждения родителя",
            COL_STATUS: "Статус_родителя",
        }, inplace=True)

        if COL_EXECUTOR in parents_best.columns:
            parents_best.rename(columns={COL_EXECUTOR: "Исполнитель_родителя"}, inplace=True)

        # merge: подтянуть к досъему данные родителя
        merge_cols = ["_p3", "№ Родительского заказа", "Дата утверждения родителя", "Статус_родителя"]
        if "Исполнитель_родителя" in parents_best.columns:
            merge_cols.append("Исполнитель_родителя")

        dos_join = dos.merge(
            parents_best[merge_cols],
            on="_p3",
            how="left"
        )

        # условие: родитель найден/утвержден, досъем НЕ утвержден (нет даты МГГТ)
        need_mask = dos_join["№ Родительского заказа"].notna() & dos_join[COL_ASU_APPROVE].isna()
        need = dos_join.loc[need_mask].copy()

        if len(need) > 0:
            dos_out = pd.DataFrame({
                "Тип проблемы": "Досъем без утверждения при утвержденном родителе",

                # от ДОСЪЁМА
                COL_ORDER: need[COL_ORDER],
                COL_ASU_APPROVE: need[COL_ASU_APPROVE],     # пустая (нормально)
                COL_ASU_LOAD_ASD: need[COL_ASU_LOAD_ASD],   # от досъёма

                # от РОДИТЕЛЯ (отдельные колонки)
                "№ Родительского заказа": need["№ Родительского заказа"],
                "Дата утверждения родителя": need["Дата утверждения родителя"],
                "Статус родителя": need["Статус_родителя"],

                # в итоговые колонки тоже кладём РОДИТЕЛЯ
                COL_EXECUTOR: need["Исполнитель_родителя"] if "Исполнитель_родителя" in need.columns else pd.NA,
                COL_STATUS: need["Статус_родителя"],
            })

    #  Склейка
    frames = []
    if len(dos_out) > 0:
        frames.append(dos_out)
    if len(bad_dates_out) > 0:
        frames.append(bad_dates_out)

    if not frames:
        return pd.DataFrame()

    out = pd.concat(frames, ignore_index=True)

    # порядок колонок
    cols_order = [
        "Тип проблемы",
        COL_ORDER,
        "№ Родительского заказа",
        "Дата утверждения родителя",
        "Статус родителя",
        COL_ASU_APPROVE,
        COL_ASU_LOAD_ASD,
        COL_EXECUTOR,
        COL_STATUS,
    ]
    out = out.loc[:, [c for c in cols_order if c in out.columns]].copy()

    return out

