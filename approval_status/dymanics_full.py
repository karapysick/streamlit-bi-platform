import pandas as pd
from datetime import datetime, timedelta
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.cell import MergedCell
from openpyxl.drawing.image import Image as XLImage
import re
from typing import Optional, Tuple
import os
import tempfile
import math
import logging

try:
    import matplotlib.pyplot as plt
except Exception:
    plt = None



# УТИЛИТЫ

def truncate_float(value, decimals=1):
    if pd.isna(value) or not isinstance(value, (int, float)):
        return value
    m = 10 ** decimals
    return int(value * m) / m


def normalize_text(value) -> str:
    s = "" if value is None else str(value)

    s = (s.replace("\u00A0", " ")
           .replace("\u202F", " ")
           .replace("\u2009", " ")
           .replace("\t", " ")
           .replace("\r", " ")
           .replace("\n", " "))


    s = (s.replace("a", "а").replace("A", "А")
           .replace("e", "е").replace("E", "Е")
           .replace("o", "о").replace("O", "О")
           .replace("p", "р").replace("P", "Р")
           .replace("c", "с").replace("C", "С")
           .replace("x", "х").replace("X", "Х")
           .replace("y", "у").replace("Y", "У"))

    s = s.replace("ё", "е").replace("Ё", "Е")
    s = re.sub(r"\s+", " ", s)
    return s.strip().lower()


def read_excel_flexible(path: str, sheet_name=0) -> pd.DataFrame:
    candidates = []
    for hdr in (2, 0):
        try:
            df = pd.read_excel(path, sheet_name=sheet_name, header=hdr)
            cols_norm = [normalize_text(c) for c in df.columns]
            score = 0
            if any("га" in c for c in cols_norm):
                score += 2
            if any(("объем" in c or "объём" in c) for c in cols_norm):
                score += 1
            if any("тип объекта огх" in c for c in cols_norm):
                score += 1
            if any("ген. договор" in c or "номер дог" in c for c in cols_norm):
                score += 1
            candidates.append((score, hdr, df))
        except Exception:
            continue

    if not candidates:
        return pd.read_excel(path, sheet_name=sheet_name)

    candidates.sort(key=lambda x: (x[0], 1 if x[1] == 2 else 0), reverse=True)
    return candidates[0][2]

def get_reports_dir() -> str:
    reports_dir = "Готовые отчеты"
    if os.path.isdir(reports_dir):
        return reports_dir
    return ""

# БЕЗОПАСНАЯ РАБОТА С MERGE

class HeaderBuilder:
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    @staticmethod
    def _find_merge_covering(ws, row, col):
        for mr in ws.merged_cells.ranges:
            c1, r1, c2, r2 = range_boundaries(str(mr))
            if r1 <= row <= r2 and c1 <= col <= c2:
                return (r1, c1, r2, c2)
        return None

    @staticmethod
    def _unmerge_point(ws, r, c):
        for mr in list(ws.merged_cells.ranges):
            c1, r1, c2, r2 = range_boundaries(str(mr))
            if r1 <= r <= r2 and c1 <= c <= c2:
                ws.unmerge_cells(str(mr))
                break

    @staticmethod
    def set_cell_value_safe(ws, row, col, value):
        rng = HeaderBuilder._find_merge_covering(ws, row, col)
        if rng:
            r1, c1, _, _ = rng
            anchor = ws.cell(row=r1, column=c1)
            anchor.value = value
            return anchor
        cell = ws.cell(row=row, column=col)
        cell.value = value
        return cell

    @staticmethod
    def _apply_range(ws, r1, c1, r2, c2, font=None, fill=None, align=None, border=None):
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell, MergedCell):
                    continue
                if font is not None:
                    cell.font = font
                if fill is not None:
                    cell.fill = fill
                if align is not None:
                    cell.alignment = align
                if border is not None:
                    cell.border = border

    @staticmethod
    def _unmerge_overlaps(ws, r1, c1, r2, c2):
        for mr in list(ws.merged_cells.ranges):
            bc1, br1, bc2, br2 = range_boundaries(str(mr))
            if not (r2 < br1 or br2 < r1 or c2 < bc1 or bc2 < c1):
                ws.unmerge_cells(str(mr))

    @staticmethod
    def set_merged(ws, r1, c1, r2, c2, value, font=None, fill=None, align=None, border=None):
        HeaderBuilder._unmerge_overlaps(ws, r1, c1, r2, c2)
        HeaderBuilder._unmerge_point(ws, r1, c1)
        anchor = ws.cell(row=r1, column=c1)
        anchor.value = value
        if r1 != r2 or c1 != c2:
            ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        HeaderBuilder._apply_range(ws, r1, c1, r2, c2, font, fill, align, border)


# ГРАНИЦЫ / РАЗДЕЛИТЕЛИ


THICK = Side(style="thick")

def _set_left_thick(ws, row, col):
    cell = ws.cell(row=row, column=col)
    b = cell.border or Border()
    cell.border = Border(left=THICK, right=b.right, top=b.top, bottom=b.bottom)


def _set_right_thick(ws, row, col):
    cell = ws.cell(row=row, column=col)
    b = cell.border or Border()
    cell.border = Border(left=b.left, right=THICK, top=b.top, bottom=b.bottom)


def _set_top_thick(ws, row, col):
    cell = ws.cell(row=row, column=col)
    b = cell.border or Border()
    cell.border = Border(left=b.left, right=b.right, top=THICK, bottom=b.bottom)


def _set_bottom_thick(ws, row, col):
    cell = ws.cell(row=row, column=col)
    b = cell.border or Border()
    cell.border = Border(left=b.left, right=b.right, top=b.top, bottom=THICK)


def draw_outline(ws, row_start, col_start, row_end, col_end):
    for c in range(col_start, col_end + 1):
        _set_top_thick(ws, row_start, c)
        _set_bottom_thick(ws, row_end, c)
    for r in range(row_start, row_end + 1):
        _set_left_thick(ws, r, col_start)
        _set_right_thick(ws, r, col_end)


def draw_vertical_divider(ws, col, row_start, row_end, left=True, style="thin"):
    side = Side(style=style)
    for r in range(row_start, row_end + 1):
        cell = ws.cell(row=r, column=col)
        b = cell.border or Border()
        if left:
            cell.border = Border(left=side, right=b.right, top=b.top, bottom=b.bottom)
        else:
            cell.border = Border(left=b.left, right=side, top=b.top, bottom=b.bottom)


def draw_horizontal_divider(ws, row, col_start, col_end, style="thick"):
    side = Side(style=style)
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        b = cell.border or Border()
        cell.border = Border(left=b.left, right=b.right, top=b.top, bottom=side)
        cell2 = ws.cell(row=row + 1, column=c)
        b2 = cell2.border or Border()
        cell2.border = Border(left=b2.left, right=b2.right, top=side, bottom=b2.bottom)


def _merge_with_thin_preserving(b: Optional[Border]) -> Border:
    thin = Side(style="thin")
    b = b or Border()

    def keep(old, fallback):
        return old if (old and getattr(old, "style", None)) else fallback

    return Border(
        left=keep(b.left, thin),
        right=keep(b.right, thin),
        top=keep(b.top, thin),
        bottom=keep(b.bottom, thin),
    )


# ГЕНЕРАТОР


class ReportGenerator:
    OP_LOAD = " Перезагрузка в АСУ ОДС\nпо дате интеграции"
    OP_REJECT = " Отклонение в АСУ ОДС\nпо дате интеграции"
    OP_APPROVE = " Утверждение в АСУ ОДС\nпо сводкам АСД"

    def __init__(self, file_path: str):
        if plt is None:
            raise RuntimeError("❌ Не установлен matplotlib. Установи: pip install matplotlib")

        self.df_selection = read_excel_flexible(file_path, sheet_name="Sheet1")
        self.df_selection = self.df_selection[[c for c in self.df_selection.columns if c not in ("Ссылка", "Примечание")]]

        # Колонка с договором
        self.contract_col = None
        for col in self.df_selection.columns:
            s = normalize_text(col)
            if any(x in s for x in ["ген. договор", "ген договор", "номер дог", "№ ген. договора", "ген.договор"]):
                self.contract_col = str(col).strip()
                break
        if not self.contract_col:
            raise KeyError(f"❌ Не найден столбец 'Ген. договор'. Колонки: {list(self.df_selection.columns)}")

        # Тип ОГХ
        self.type_col = None
        for col in self.df_selection.columns:
            if "тип объекта огх" in normalize_text(col):
                self.type_col = str(col).strip()
                break
        if not self.type_col:
            raise KeyError("❌ Не найден столбец 'Тип объекта ОГХ'.")

        order_col = "№ Заказа МГГТ"

        status_col = None
        for col in self.df_selection.columns:
            s = normalize_text(col)
            if ("состояние" in s) or ("статус" in s and "аннулир" in s):
                status_col = str(col).strip()
                break

        #  РВК
        if order_col in self.df_selection.columns:
            has_RVK = self.df_selection[order_col].astype(str).str.contains(r"[РКВЮ]", na=False, case=False)
        else:
            has_RVK = pd.Series(False, index=self.df_selection.index)

        #  действующий и приостановленный
        if status_col and status_col in self.df_selection.columns:
            st = self.df_selection[status_col].astype(str).str.lower()
            is_active = st.str.contains("действ", na=False) | st.str.contains("приостанов", na=False)
        else:
            logger = logging.getLogger(__name__)
            logger.warning("Не найдена колонка статуса. Аннулированные записи могут не отфильтроваться.")
            is_active = pd.Series(True, index=self.df_selection.index)

        # Итоговый фильтр
        self.selection_filtered = self.df_selection[(~has_RVK) & (is_active)].copy()

        # Нормализованный тип ОГХ
        TYPE_NORM = "_TYPE_NORM"
        self.selection_filtered[TYPE_NORM] = (
            self.selection_filtered[self.type_col]
            .astype(str)
            .str.replace("\u00A0", " ", regex=False)
            .str.strip()
            .str.replace(r"\s+", "", regex=True)
            .str.upper()
        )
        self.type_norm_col = TYPE_NORM

        # 3 операции
        self.operation_map = {
            self.OP_LOAD: "Дата загрузки в АСУ ОДС",
            self.OP_REJECT: "Дата отклонения в АСУ ОДС",
            self.OP_APPROVE: "Дата утверждения в АСУ ОДС (МГГТ)",
        }
        missing = [col for col in self.operation_map.values() if col not in self.selection_filtered.columns]
        if missing:
            raise KeyError("❌ В БД не найдены колонки:\n- " + "\n- ".join(missing))

        for col in self.operation_map.values():
            self.selection_filtered[col] = pd.to_datetime(self.selection_filtered[col], errors="coerce", dayfirst=True)

        self.operation_order = [self.OP_LOAD, self.OP_REJECT, self.OP_APPROVE]

        # заливки
        base_color = "FCD5B4"
        self.operation_fills = {op: PatternFill(start_color=base_color, end_color=base_color, fill_type="solid")
                                for op in self.operation_order}

        self.fact_fill = PatternFill(start_color="C6E0B4", fill_type="solid")
        self.percent_fill = PatternFill(start_color="B4C6E7", fill_type="solid")
        self.work_fill = PatternFill(start_color="FFFF99", fill_type="solid")
        self.subgroup_fill = PatternFill(start_color="FFE6CC", fill_type="solid")

        self.left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        self.center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.center_nowrap = Alignment(horizontal="center", vertical="center", wrap_text=False)
        self.bold = Font(bold=True)
        self.normal = Font(bold=False)
        self.thin = HeaderBuilder.thin

        self.remaining_fill = PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid")
        self.remaining_font = Font(color="9C0006", bold=True)
        self.ogx_type_fill = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")

        # подписи линий
        self.chart_series_titles = {
            self.OP_LOAD: "Перезагрузка в АСУ ОДС по дате интеграции",
            self.OP_REJECT: "Отклонение в АСу ОДС по дате интеграции",
            self.OP_APPROVE: "Утверждение в АСУ ОДС по сводкам АСД",
        }

        # цвета линий
        self.chart_line_colors = {
            self.OP_LOAD: "#4472C4",
            self.OP_REJECT: "#C00000",
            self.OP_APPROVE: "#00B050",
        }

    def infer_gz_year(self) -> int:
        s = self.selection_filtered[self.contract_col].astype(str)
        m = s.str.extract(r"ОГХ[-\s]?(\d{2,4})", flags=re.IGNORECASE)[0].dropna()
        if m.empty:
            return int(datetime.now().strftime("%y"))

        vals = []
        for x in m.tolist():
            x = str(x).strip()
            if len(x) == 4:
                vals.append(int(x[-2:]))
            elif len(x) == 2:
                vals.append(int(x))
        if not vals:
            return int(datetime.now().strftime("%y"))

        return int(pd.Series(vals).mode().iloc[0])

    def filter_by_gz_year(self, df, gz_year: int):
        patterns = [
            f"12/ОГХ-{gz_year}",
            f"12/ОГХ-{2000 + gz_year}",
            f"ОГХ-{gz_year}",
            f"ОГХ-{2000 + gz_year}",
        ]
        col_data = df[self.contract_col].astype(str).str.lower()
        mask = pd.Series(False, index=df.index)
        for p in patterns:
            mask |= col_data.str.contains(p.lower(), na=False)
        return df[mask].copy()

    def _build_header(self, ws, periods):
        top = 2
        ws.row_dimensions[top].height = 40
        ws.row_dimensions[top + 1].height = 22

        HeaderBuilder.set_merged(ws, top, 1, top + 1, 1,
                                 "Стадия выполнения/Период",
                                 font=self.bold, border=self.thin, align=self.center)
        HeaderBuilder.set_merged(ws, top, 2, top + 1, 2,
                                 "ТИП ОГХ",
                                 font=self.bold, border=self.thin, align=self.center)

        first_period_col = 3

        for i, p in enumerate(periods):
            c = first_period_col + i
            cell = ws.cell(row=top, column=c)
            cell.value = p
            cell.font = self.normal
            cell.alignment = self.center_nowrap
            cell.border = self.thin

            cell2 = ws.cell(row=top + 1, column=c)
            cell2.border = self.thin
            cell2.alignment = self.center_nowrap

        fact_col = first_period_col + len(periods)
        percent_col = fact_col + 1
        remaining_col = percent_col + 1

        HeaderBuilder.set_merged(ws, top, fact_col, top + 1, fact_col,
                                 "Факт выполнения Графика", font=self.bold, fill=self.fact_fill,
                                 border=self.thin, align=self.center)
        HeaderBuilder.set_merged(ws, top, percent_col, top + 1, percent_col,
                                 "% выполнения графика", font=self.bold, fill=self.percent_fill,
                                 border=self.thin, align=self.center)
        HeaderBuilder.set_merged(ws, top, remaining_col, top + 1, remaining_col,
                                 "Осталось выполнить", font=self.bold, fill=self.remaining_fill,
                                 border=self.thin, align=self.center)

        return top + 2, first_period_col, fact_col, percent_col, remaining_col

    def _auto_width(self, ws):
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) for cell in col
                 if cell.value is not None and not isinstance(cell, MergedCell)),
                default=0
            )
            width = max(10, min(max_len + 2, 60))
            ws.column_dimensions[get_column_letter(col[0].column)].width = width

        ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 56)
        ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 0, 18)

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                cell.border = _merge_with_thin_preserving(cell.border)

    def _unify_row_heights(self, ws, start_row=3, default_h=22):
        for r in range(start_row, ws.max_row + 1):
            ws.row_dimensions[r].height = default_h

    def _iter_rows_for_daily(self, df):
        def mask_total():
            return pd.Series(True, index=df.index)

        type_series = df[self.type_norm_col].dropna().astype(str)
        dyn_types = sorted([x for x in type_series.unique() if x], key=str)
        preferred = ["ДТ", "ОО", "ОДХ"]
        ogx_types = preferred + [t for t in dyn_types if t not in preferred]

        def make_mask_by_type(type_val):
            tv = str(type_val).replace("\u00A0", " ").strip()
            tv = "".join(tv.split()).upper()
            return lambda: (df[self.type_norm_col] == tv)

        for t in ogx_types:
            yield make_mask_by_type(t), [t]

        yield mask_total, ["ВСЕГО"]


    # ГРАФИК + ТАБЛИЦА

    def _make_matplotlib_chart_png_with_table(
            self,
            x_labels,
            series_rows_in_order,
            out_png_path,
            width_cols_span: int,
            is_sht: bool,
    ) -> Tuple[int, int]:
        dpi = 150

        # Ширина графика
        fig_w = max(18.0, min(40.0, width_cols_span * 0.85))
        fig_h = 5.6

        # ДВА AXES: сверху график, снизу таблица
        fig, (ax, ax_tbl) = plt.subplots(
            nrows=2,
            ncols=1,
            figsize=(fig_w, fig_h),
            dpi=dpi,
            gridspec_kw={"height_ratios": [4.2, 1.0], "hspace": 0.08}
        )

        x = list(range(len(x_labels)))
        x_shifted = [i - 0.07 for i in x]

        for row_label, y_list, color in series_rows_in_order:
            ax.plot(x, y_list, linewidth=2.2, label=row_label, color=color)

        # ось X: подписи
        ax.set_xticks(x_shifted)
        ax.set_xticklabels(x_labels, rotation=0, fontsize=8)

        # обрезка
        ax.set_xlim(-0.5, len(x_labels) - 0.5)

        # ось Y - авто под данные
        ymax_raw = max([max(y) for _, y, _ in series_rows_in_order] + [0])

        #  запас сверху
        ymax_raw = max(1, float(ymax_raw) * 1.15)

        # округление для красоты

        if ymax_raw <= 200:
            step = 25
        elif ymax_raw <= 500:
            step = 50
        elif ymax_raw <= 1000:
            step = 100
        elif ymax_raw <= 2000:
            step = 200
        else:
            step = 500

        ymax = int(math.ceil(ymax_raw / step) * step)
        ax.set_ylim(0, ymax)

        ax.set_yticks(list(range(0, ymax + 1, step)))

        # сетка
        ax.grid(True, axis="y", linewidth=0.8)
        ax.grid(False, axis="x")

        # легенда
        ax.legend(loc="center left", bbox_to_anchor=(1.01, 0.5), frameon=False, fontsize=8)

        # таблица снизу
        ax_tbl.axis("off")

        cell_text = []
        row_labels = []
        for row_label, y_list, _ in series_rows_in_order:
            row_labels.append(row_label)
            cell_text.append([int(v) if v is not None else 0 for v in y_list])

        # Таблица БЕЗ дат
        tbl = ax_tbl.table(
            cellText=cell_text,
            rowLabels=row_labels,
            cellLoc="center",
            rowLoc="center",
            loc="center"
        )
        tbl.auto_set_font_size(False)
        tbl.set_fontsize(7)
        tbl.scale(1.0, 1.15)

        # расширяем область таблицы
        pos = ax_tbl.get_position()

        ax_tbl.set_position([
            max(pos.x0 - 0.05, 0.02),
            pos.y0,
            pos.width + 0.05,
            pos.height
        ])

        # место справа под Малые доли
        fig.subplots_adjust(left=0.24, right=0.62, top=0.95, bottom=0.07)

        # сохраняем с запасом
        fig.savefig(out_png_path, bbox_inches="tight", pad_inches=0.35)
        plt.close(fig)

        return int(fig_w * dpi), int(fig_h * dpi)


    def _add_matplotlib_chart_block(
        self,
        ws,
        sheet_kind: str,
        base_col: int,
        fact_col: int,
        periods,
        total_rows_by_op,
        ops_order,
        png_prefix: str
    ):
        """
        Вставляет ОДНУ картинку: график + таблица значений внутри картинки.
        Якорь по ширине: начинаем с первой колонки дат (base_col),
        заканчиваем на fact_col (включительно).
        """
        n_periods = len(periods)

        series_rows = []
        for op in ops_order:
            r = total_rows_by_op.get(op)
            y = []
            if r:
                for j in range(n_periods):
                    cell_value = ws.cell(row=r, column=base_col + j).value
                    try:
                        y.append(int(cell_value) if cell_value not in (None, "") else 0)
                    except Exception:
                        y.append(0)
            else:
                y = [0] * n_periods

            label = self.chart_series_titles.get(op, op)
            color = self.chart_line_colors.get(op, "#000000")
            series_rows.append((label, y, color))

        width_cols_span = fact_col

        stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        png_name = f"{png_prefix}_{stamp}.png"
        reports_dir = get_reports_dir()
        png_path = os.path.join(reports_dir, png_name) if reports_dir else png_name

        is_sht = (sheet_kind == "Шт")
        px_w, px_h = self._make_matplotlib_chart_png_with_table(
            x_labels=periods,
            series_rows_in_order=series_rows,
            out_png_path=png_path,
            width_cols_span=width_cols_span,
            is_sht=is_sht
        )

        # вставка ниже таблицы
        anchor_row = ws.max_row + 1
        anchor_cell = f"A{anchor_row}"

        img = XLImage(png_path)

        img.width = int(px_w * 1.10)
        img.height = int(px_h * 0.88)

        ws.add_image(img, anchor_cell)

    def _add_matplotlib_chart_block_workdays(
            self,
            ws,
            sheet_kind: str,
            base_col: int,
            fact_col: int,
            periods_full,
            workday_mask,
            total_rows_by_op,
            ops_order,
            png_prefix: str
    ):
        """
        То же самое что _add_matplotlib_chart_block, но:
        - X_labels = только будни
        - значения Y берём из таблицы по всем дням, но фильтруем по workday_mask
        """
        n_full = len(periods_full)
        idx_work = [i for i, ok in enumerate(workday_mask) if ok]
        periods_work = [periods_full[i] for i in idx_work]

        series_rows = []
        for op in ops_order:
            r = total_rows_by_op.get(op)
            y_full = []
            if r:
                for j in range(n_full):
                    cell_value = ws.cell(row=r, column=base_col + j).value
                    try:
                        y_full.append(int(cell_value) if cell_value not in (None, "") else 0)
                    except Exception:
                        y_full.append(0)
            else:
                y_full = [0] * n_full

            # фильтруем по будням
            y_work = [y_full[i] for i in idx_work]

            label = self.chart_series_titles.get(op, op)
            color = self.chart_line_colors.get(op, "#000000")
            series_rows.append((label, y_work, color))

        width_cols_span = fact_col

        stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        png_name = f"{png_prefix}_{stamp}.png"
        reports_dir = get_reports_dir()
        png_path = os.path.join(reports_dir, png_name) if reports_dir else png_name

        is_sht = (sheet_kind == "Шт")
        px_w, px_h = self._make_matplotlib_chart_png_with_table(
            x_labels=periods_work,
            series_rows_in_order=series_rows,
            out_png_path=png_path,
            width_cols_span=width_cols_span,
            is_sht=is_sht
        )

        anchor_row = ws.max_row + 1
        anchor_cell = f"A{anchor_row}"

        img = XLImage(png_path)
        img.width = int(px_w * 1.10)
        img.height = int(px_h * 0.88)
        ws.add_image(img, anchor_cell)

    def _clear_cell_no_grid(self, cell):
        cell.value = None
        cell.border = Border()
        cell.fill = PatternFill()


    # ОСНОВНОЙ ОТЧЕТ


    def generate_daily_ops_split_ogx(self, gz_year_override: Optional[int] = None):
        gz_year = int(gz_year_override) if gz_year_override is not None else self.infer_gz_year()
        df = self.filter_by_gz_year(self.selection_filtered, gz_year)

        today = datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)
        weekday = today.weekday()
        end_date = today - timedelta(days=1) if weekday == 0 else (today - timedelta(days=weekday) + timedelta(days=6))
        start_date = end_date - timedelta(days=20)
        date_range = pd.date_range(start=start_date, end=end_date, freq="D")

        months_ru = ["янв", "фев", "мар", "апр", "май", "июн", "июл", "авг", "сен", "окт", "ноя", "дек"]
        periods = [f"{d.day} {months_ru[d.month - 1]}" for d in date_range]

        # для таблицы: все дни
        date_range_full = date_range
        periods_full = periods

        # для графика: только будни
        workday_mask = [d.weekday() < 5 for d in date_range_full]  # 0..4 = будни
        date_range_work = [d for d, ok in zip(date_range_full, workday_mask) if ok]
        periods_work = [p for p, ok in zip(periods_full, workday_mask) if ok]

        # для графика: показываем только до ВЧЕРА
        yesterday = today - timedelta(days=1)

        idx_yesterday = [i for i, d in enumerate(date_range_full) if d <= yesterday]
        last_idx = max(idx_yesterday) if idx_yesterday else len(date_range_full) - 1

        periods_full_chart = periods_full[:last_idx + 1]
        workday_mask_chart = workday_mask[:last_idx + 1]

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"ежедневный_АВТО_отчет_интеграция_ГЗ-{gz_year}_{timestamp}.xlsx"

        reports_dir = get_reports_dir()
        out_path = os.path.join(reports_dir, filename) if reports_dir else filename

        rows_def = list(self._iter_rows_for_daily(df))

        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            wb = writer.book
            ws_count = wb.create_sheet("Шт.")

            #  ШАПКА ОТЧЁТА над таблицей
            first_period_col = 3
            fact_col_tmp = first_period_col + len(periods)
            percent_col_tmp = fact_col_tmp + 1
            remain_col_tmp = percent_col_tmp + 1

            report_date_str = datetime.now().strftime("%d.%m.%Y")
            title_text = f"Ежедневный_{gz_year} отчёт по паспортизации от {report_date_str}"

            ws_count.row_dimensions[1].height = 28
            HeaderBuilder.set_merged(
                ws_count, 1, 1, 1, remain_col_tmp,
                title_text,
                font=Font(bold=True, size=14),
                align=self.center,
                border=self.thin
            )

            row_c, base_c, fact_c, percent_c, remain_c = self._build_header(ws_count, periods)
            header_bottom_c = row_c - 1

            row_count = row_c

            prev_fact_count_by_labels = {}

            total_row_count_by_op = {}

            for op in self.operation_order:
                col_name = self.operation_map[op]
                blank_summary = (op in (self.OP_LOAD, self.OP_REJECT))

                # толстая линия перед блоком операции
                draw_horizontal_divider(ws_count, row_count - 1, 1, remain_c, style="thick")

                # считаем факт по маскам заранее
                current_fact_count = {}
                for mask_func, labels in rows_def:
                    mask = mask_func()
                    fact_mask = mask & df[col_name].notna()
                    current_fact_count[tuple(labels)] = int(fact_mask.sum())

                # рисуем 4 строки (ВСЕГО/ДТ/ОО/ОДХ)
                row_count_temp = row_count
                block_start_row = row_count_temp

                for mask_func, labels in rows_def:
                    mask = mask_func()
                    is_total = (len(labels) > 0 and labels[0] == "ВСЕГО")
                    ogx_label = labels[0] if labels else ""

                    # Колонка B: тип ОГХ
                    is_ogx_type = ogx_label in ("ДТ", "ОО", "ОДХ")
                    HeaderBuilder.set_merged(
                        ws_count, row_count_temp, 2, row_count_temp, 2, ogx_label,
                        font=self.bold if is_total else self.normal,
                        fill=(
                            self.subgroup_fill if is_total else
                            self.ogx_type_fill if is_ogx_type else
                            None
                        ),
                        align=self.center,
                        border=self.thin
                    )

                    # Колонки дат
                    for j, date in enumerate(date_range):
                        date_mask = mask & (df[col_name].dt.date == date.date())
                        cnt = int(date_mask.sum())

                        cell_cnt = HeaderBuilder.set_cell_value_safe(ws_count, row_count_temp, base_c + j, cnt)
                        cell_cnt.alignment = self.center
                        cell_cnt.border = self.thin
                        cell_cnt.fill = self.operation_fills[op]
                        if is_total:
                            cell_cnt.font = self.bold

                    # правые колонки: факт / % / осталось
                    key = tuple(labels)
                    fact_count = current_fact_count.get(key, 0)
                    #  ПЛАН для текущей строки (ВСЕГО / ДТ / ОО / ОДХ)
                    plan_row = int(mask.sum())

                    percent_value = (fact_count / plan_row) if plan_row > 0 else 0
                    remaining_count = max(plan_row - fact_count, 0)

                    if not blank_summary:
                        c = HeaderBuilder.set_cell_value_safe(ws_count, row_count_temp, fact_c, fact_count)
                        c.font = self.bold
                        c.fill = self.fact_fill
                        c.alignment = self.center
                        c.border = self.thin

                        c = HeaderBuilder.set_cell_value_safe(ws_count, row_count_temp, percent_c, percent_value)
                        c.font = self.bold
                        c.fill = self.percent_fill
                        c.alignment = self.center
                        c.border = self.thin
                        c.number_format = "0.0%"

                        c = HeaderBuilder.set_cell_value_safe(ws_count, row_count_temp, remain_c, remaining_count)
                        c.font = self.remaining_font
                        c.fill = self.remaining_fill
                        c.alignment = self.center
                        c.border = self.thin

                    if is_total:
                        total_row_count_by_op[op] = row_count_temp

                    row_count_temp += 1

                block_end_row = row_count_temp - 1

                if blank_summary:
                    for col in (fact_c, percent_c, remain_c):
                        for r in range(block_start_row, block_end_row + 1):
                            cell = ws_count.cell(row=r, column=col)
                            if isinstance(cell, MergedCell):
                                continue
                            cell.value = None
                            cell.border = Border()
                            cell.fill = PatternFill()
                            cell.alignment = self.center

                        ws_count.merge_cells(
                            start_row=block_start_row, start_column=col,
                            end_row=block_end_row, end_column=col
                        )

                        tl = ws_count.cell(row=block_start_row, column=col)
                        tl.value = None
                        tl.border = self.thin
                        tl.fill = PatternFill()
                        tl.alignment = self.center

                # название операции
                HeaderBuilder.set_merged(
                    ws_count, block_start_row, 1, block_end_row, 1, op,
                    font=self.bold,
                    fill=self.operation_fills[op],
                    align=self.left,
                    border=self.thin
                )

                # толстая линия после блока операции
                draw_horizontal_divider(ws_count, block_end_row, 1, remain_c, style="thick")

                # следующий блок - через одну пустую строку
                row_count = block_end_row + 2

                draw_horizontal_divider(ws_count, block_end_row, 1, remain_c, style="thick")

                row_count = row_count_temp + 1

                prev_fact_count_by_labels = current_fact_count

            draw_horizontal_divider(ws_count, header_bottom_c, 1, remain_c, style="thick")

            draw_vertical_divider(ws_count, 2, 1, ws_count.max_row, left=True, style="thin")

            draw_vertical_divider(ws_count, base_c, 1, ws_count.max_row, left=True, style="thin")

            for col in (fact_c, percent_c, remain_c):
                draw_vertical_divider(ws_count, col, 1, ws_count.max_row, left=True, style="thin")

            draw_outline(ws_count, 1, 1, ws_count.max_row, remain_c)

            self._auto_width(ws_count)
            # ФИКСИРОВАННАЯ ширина КОЛОНОК ДАТ в ОСНОВНОЙ ТАБЛИЦЕ
            for j in range(len(periods)):
                ws_count.column_dimensions[get_column_letter(base_c + j)].width = 15.85

            # ГРАФИКИ (картинка): старт = первая дата, конец = колонка Факт выполнения графика
            self._add_matplotlib_chart_block_workdays(
                ws=ws_count,
                sheet_kind="Шт",
                base_col=base_c,
                fact_col=fact_c,
                periods_full=periods_full_chart,
                workday_mask=workday_mask_chart,
                total_rows_by_op=total_row_count_by_op,
                ops_order=self.operation_order,
                png_prefix="chart_sht"
            )

            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        return gz_year, out_path

# STREAMLIT WRAPPER

def build_daily_report_streamlit(
    db_excel_bytes: bytes,
    gz_year: Optional[int] = None,
) -> Tuple[bytes, str, int]:
    """
    Обёртка для Streamlit:
    - принимает Excel БД в bytes
    - использует ReportGenerator для формирования отчёта
    - возвращает (excel_bytes, filename, gz_year_real)
    """

    if not db_excel_bytes:
        raise ValueError("Пустой Excel-файл БД (db_excel_bytes).")

    #  bytes -> временный xlsx
    tmp_in = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp_in.write(db_excel_bytes)
    tmp_in.close()

    out_path = None

    try:
        # генератор
        generator = ReportGenerator(tmp_in.name)
        gz_year_real, out_path = generator.generate_daily_ops_split_ogx(
            gz_year_override=gz_year
        )

        #  читаем результат в bytes
        with open(out_path, "rb") as f:
            excel_bytes = f.read()

        filename = f"Статус_утверждения_{datetime.now().strftime('%d.%m.%Y_%H-%M')}.xlsx"
        return excel_bytes, filename, int(gz_year_real)

    finally:
        try:
            os.remove(tmp_in.name)
        except Exception:
            pass

        # чистка
        try:
            if out_path and os.path.exists(out_path):
                os.remove(out_path)
        except Exception:
            pass