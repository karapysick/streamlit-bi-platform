# full_report.py
# Файл-склейка dymanics_full.py + everyday_1.py
# ✅ Подготовлено для Streamlit: НЕ открывает Excel, умеет возвращать bytes

from __future__ import annotations

import os
import tempfile
import datetime as dt
from typing import Optional, Tuple

import pandas as pd
from openpyxl import load_workbook

from . import everyday_1 as everyday
from .Github import ReportGenerator


def build_full_report(db_path: str, gz_year: Optional[int] = None, open_file: bool = False) -> Tuple[str, pd.DataFrame, int]:
    """
    Собирает полный отчёт на диске (xlsx) и возвращает:
      - out_path: путь к сформированному xlsx
      - issues_df: таблица проблемных заказов (может быть пустой DataFrame)
      - filtered_orders_count: кол-во строк после фильтрации
    """

    out_path: Optional[str] = None
    issues_df: pd.DataFrame = pd.DataFrame()
    filtered_orders_count: int = 0

    #  1. dymanics_full.py
    generator = ReportGenerator(db_path)
    gz_year_real, out_path = generator.generate_daily_ops_split_ogx(
        gz_year_override=gz_year
    )

    # 2. Открываем workbook
    wb = load_workbook(out_path)
    ws = wb["Шт."]

    def find_last_table_row(_ws) -> int:
        last = 1
        for r in range(1, _ws.max_row + 1):
            v = _ws.cell(r, 1).value
            if isinstance(v, str) and v.strip():
                last = r
        return last

    #  3. Где начинаем everyday
    github_last_table_row = find_last_table_row(ws)
    start_row = github_last_table_row + 43

    # 4. Фильтры (один раз)
    df = everyday.load_db_excel(db_path)
    dfp, _ = everyday.prepare_filtered_df(df, gz_year_real)
    issues_df = everyday.build_closure_issues(dfp)
    filtered_orders_count = len(dfp)

    approved_mask, rejected_mask, review_mask = everyday.build_masks(dfp)
    remain_mask = ~approved_mask

    # 5. EVERYDAY — ШТ
    total_c, appr_c, rej_c, rev_c, rem_c = everyday.aggregate_counts(
        dfp, approved_mask, rejected_mask, review_mask
    )

    last_row = everyday.write_table(
        ws=ws,
        start_row=start_row,
        start_col=2,
        title_year_suffix=gz_year_real,
        report_date=dt.date.today(),
        unit_label="ШТ",
        total=total_c,
        appr=appr_c,
        rej=rej_c,
        rev=rev_c,
        remain=rem_c,
        is_float=False,
        round_ga=False,
    )

    #  6. EVERYDAY — ГА
    total_h, appr_h, rej_h, rev_h, rem_h = everyday.aggregate_ha(
        dfp, approved_mask, rejected_mask, review_mask
    )

    last_row_2 = everyday.write_table(
        ws=ws,
        start_row=last_row + 4,
        start_col=2,
        title_year_suffix=gz_year_real,
        report_date=dt.date.today(),
        unit_label="ГА",
        total=total_h,
        appr=appr_h,
        rej=rej_h,
        rev=rev_h,
        remain=rem_h,
        is_float=True,
        round_ga=True,
    )

    # ✅ один круговой график по ШТ, но вставляем ПОСЛЕ ГА
    ok_pie, pie_png_path = everyday.insert_status_pie_to_ws(
        ws=ws,
        anchor_row=last_row_2 + 2,
        anchor_col=1,
        year_suffix=gz_year_real,
        approved=float(appr_c.sum()),
        rejected=float(rej_c.sum()),
        review=float(rev_c.sum()),
    )

    #  7. Остальные листы
    everyday.write_df_sheet(wb, "Утверждено", dfp.loc[approved_mask], everyday.EXPORT_COLS)
    everyday.write_df_sheet(wb, "Отклонено", dfp.loc[rejected_mask], everyday.EXPORT_COLS)
    everyday.write_df_sheet(wb, "На рассмотрении", dfp.loc[review_mask], everyday.EXPORT_COLS)
    everyday.write_df_sheet(wb, "Осталось утвердить", dfp.loc[remain_mask], everyday.EXPORT_COLS)

    #  8. Сохраняем
    wb.save(out_path)

    # чистим временную png
    try:
        if pie_png_path and os.path.exists(pie_png_path):
            os.remove(pie_png_path)
    except Exception:
        pass

    # ❗ Для Streamlit по умолчанию НЕ открываем Excel
    if open_file:
        try:
            os.startfile(out_path)  # только Windows
        except Exception:
            pass

    # ✅ ЕДИНЫЙ RETURN
    return out_path, issues_df, filtered_orders_count


#  STREAMLIT WRAPPER
def build_full_report_streamlit(
    db_excel_bytes: bytes,
    gz_year: Optional[int] = None
) -> Tuple[bytes, str, pd.DataFrame, int]:
    """
    Streamlit-обёртка:
      - вход: Excel-файл БД в bytes
      - выход: (excel_bytes, filename, issues_df, filtered_orders_count)

    Внутри:
      bytes -> временный .xlsx -> build_full_report() -> читаем результат -> bytes
      затем удаляем временные файлы.
    """

    if not db_excel_bytes:
        raise ValueError("Пустой Excel-файл БД (db_excel_bytes).")

    tmp_in = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp_in.write(db_excel_bytes)
    tmp_in.close()

    out_path: Optional[str] = None

    try:
        out_path, issues_df, filtered_orders_count = build_full_report(
            db_path=tmp_in.name,
            gz_year=gz_year,
            open_file=False,  # ✅ важно для Streamlit
        )

        with open(out_path, "rb") as f:
            excel_bytes = f.read()

        filename = f"Статус_утверждения_{dt.datetime.now().strftime('%d.%m.%Y_%H-%M')}.xlsx"

        if issues_df is None:
            issues_df = pd.DataFrame()

        return excel_bytes, filename, issues_df, int(filtered_orders_count)

    finally:
        # удаляем входной временный файл
        try:
            os.remove(tmp_in.name)
        except Exception:
            pass

        # удаляем итоговый файл-результат (чтобы не мусорить)
        try:
            if out_path and os.path.exists(out_path):
                os.remove(out_path)
        except Exception:
            pass
