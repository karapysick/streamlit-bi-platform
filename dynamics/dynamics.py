import pandas as pd
from datetime import datetime, timedelta
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter, range_boundaries
import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox
import json
import re
from openpyxl.cell import MergedCell

# Утилиты

def truncate_float(value, decimals=1):
    if pd.isna(value) or not isinstance(value, (int, float)):
        return value
    m = 10 ** decimals
    return int(value * m) / m

# Нормализация строк (лечит NBSP/тонкие пробелы/табы/латиницу в кириллице и т.п.)
def normalize_text(value) -> str:
    s = "" if value is None else str(value)

    s = (s.replace("\u00A0", " ")   # NBSP
           .replace("\u202F", " ")  # narrow no-break space
           .replace("\u2009", " ")  # thin space
           .replace("\t", " ")
           .replace("\r", " ")
           .replace("\n", " "))

    # латиница -> кириллица (частые визуальные подмены)
    s = (s.replace("a", "а").replace("A", "А")
           .replace("e", "е").replace("E", "Е")
           .replace("o", "о").replace("O", "О")
           .replace("p", "р").replace("P", "Р")
           .replace("c", "с").replace("C", "С")
           .replace("x", "х").replace("X", "Х")
           .replace("y", "у").replace("Y", "У"))

    s = s.replace("ё", "е").replace("Ё", "Е")
    s = re.sub(r"\s+", " ", s)
    return s.strip().lower()


def read_excel_flexible(path: str, sheet_name=0) -> pd.DataFrame:
    """
    Читает Excel, когда шапка может быть на 1-й строке или на 3-й.
    1) Пытается header=2, затем header=0
    2) Проверяет, что среди колонок есть что-то похожее на нужные поля (например 'га')
    """
    candidates = []

    for hdr in (2, 0):
        try:
            df = pd.read_excel(path, sheet_name=sheet_name, header=hdr)
            cols_norm = [normalize_text(c) for c in df.columns]
            score = 0
            # эвристики: чем больше нашли "ключевых" признаков, тем вероятнее, что шапка правильная
            if any("га" in c for c in cols_norm):
                score += 2
            if any(("объем" in c or "объём" in c) for c in cols_norm):
                score += 1
            if any("тип объекта огх" in c for c in cols_norm):
                score += 1
            if any("ген. договор" in c or "номер дог" in c for c in cols_norm):
                score += 1
            candidates.append((score, hdr, df))
        except Exception:
            continue

    if not candidates:
        # как запасной вариант — просто обычное чтение
        return pd.read_excel(path, sheet_name=sheet_name)

    # берём лучший score, при равенстве предпочитаем header=2 (как было раньше)
    candidates.sort(key=lambda x: (x[0], 1 if x[1] == 2 else 0), reverse=True)
    return candidates[0][2]

def open_file_crossplatform(path: str):
    try:
        if sys.platform.startswith("win"):
            os.startfile(path)
        elif sys.platform.startswith("darwin"):
            os.system(f"open '{path}'")
        else:
            os.system(f"xdg-open '{path}'")
    except Exception as e:
        print(f"Не удалось открыть файл: {e}")

SETTINGS_FILE = "report_settings.json"

def load_settings():
    try:
        with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}

def save_settings(settings):
    try:
        with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
            json.dump(settings, f, ensure_ascii=False, indent=4)
    except Exception as e:
        print(f"Не удалось сохранить настройки: {e}")

app_settings = load_settings()

# Безопасная работа с объединениями

class HeaderBuilder:
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    @staticmethod
    def _find_merge_covering(ws, row, col):
        # ВАЖНО: range_boundaries возвращает (c1, r1, c2, r2)
        for mr in ws.merged_cells.ranges:
            c1, r1, c2, r2 = range_boundaries(str(mr))
            if r1 <= row <= r2 and c1 <= col <= c2:
                # Возвращаем (r1, c1, r2, c2) в удобном порядке
                return (r1, c1, r2, c2)
        return None

    @staticmethod
    def _unmerge_point(ws, r, c):
        # ВАЖНО: (c1, r1, c2, r2)
        for mr in list(ws.merged_cells.ranges):
            c1, r1, c2, r2 = range_boundaries(str(mr))
            if r1 <= r <= r2 and c1 <= c <= c2:
                ws.unmerge_cells(str(mr))
                break

    @staticmethod
    def _anchor_cell(ws, row, col):
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            rng = HeaderBuilder._find_merge_covering(ws, row, col)
            if rng:
                r1, c1, _, _ = rng
                return ws.cell(row=r1, column=c1)
        return cell

    @staticmethod
    def set_cell_value_safe(ws, row, col, value):
        """
        Пишем в якорную ячейку, если (row, col) попадает в объединение.
        Ничего не разъединяем, чтобы не ломать структуру объединённых ячеек.
        """
        rng = HeaderBuilder._find_merge_covering(ws, row, col)
        if rng:
            r1, c1, _, _ = rng
            anchor = ws.cell(row=r1, column=c1)
            anchor.value = value
            return anchor
        else:
            cell = ws.cell(row=row, column=col)
            cell.value = value
            return cell

    @staticmethod
    def _apply_range(ws, r1, c1, r2, c2, font=None, fill=None, align=None, border=None):
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell, MergedCell):
                    continue
                if font is not None:
                    cell.font = font
                if fill is not None:
                    cell.fill = fill
                if align is not None:
                    cell.alignment = align
                if border is not None:
                    cell.border = border

    @staticmethod
    def _unmerge_overlaps(ws, r1, c1, r2, c2):
        # ВАЖНО: используем (c1, r1, c2, r2) из range_boundaries
        for mr in list(ws.merged_cells.ranges):
            bc1, br1, bc2, br2 = range_boundaries(str(mr))
            if not (r2 < br1 or br2 < r1 or c2 < bc1 or bc2 < c1):
                ws.unmerge_cells(str(mr))

    @staticmethod
    def set_merged(ws, r1, c1, r2, c2, value, font=None, fill=None, align=None, border=None):
        HeaderBuilder._unmerge_overlaps(ws, r1, c1, r2, c2)
        HeaderBuilder._unmerge_point(ws, r1, c1)
        anchor = ws.cell(row=r1, column=c1)
        anchor.value = value
        if r1 != r2 or c1 != c2:
            ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        HeaderBuilder._apply_range(ws, r1, c1, r2, c2, font, fill, align, border)

# Толстые границы / рамки

THICK = Side(style="thick")

def _set_left_thick(ws, row, col):
    cell = ws.cell(row=row, column=col)
    b = cell.border or Border()
    cell.border = Border(left=THICK, right=b.right, top=b.top, bottom=b.bottom)

def _set_right_thick(ws, row, col):
    cell = ws.cell(row=row, column=col)
    b = cell.border or Border()
    cell.border = Border(left=b.left, right=THICK, top=b.top, bottom=b.bottom)

def _set_top_thick(ws, row, col):
    cell = ws.cell(row=row, column=col)
    b = cell.border or Border()
    cell.border = Border(left=b.left, right=b.right, top=THICK, bottom=b.bottom)

def _set_bottom_thick(ws, row, col):
    cell = ws.cell(row=row, column=col)
    b = cell.border or Border()
    cell.border = Border(left=b.left, right=b.right, top=b.top, bottom=THICK)

def draw_outline(ws, row_start, col_start, row_end, col_end):
    for c in range(col_start, col_end + 1):
        _set_top_thick(ws, row_start, c)
        _set_bottom_thick(ws, row_end, c)
    for r in range(row_start, row_end + 1):
        _set_left_thick(ws, r, col_start)
        _set_right_thick(ws, r, col_end)
def _merge_with_thin_preserving(b: Border | None) -> Border:
    """
    Добавляет тонкие границы там, где их нет, НЕ затирая уже заданные (в т.ч. thick).
    """
    thin = Side(style="thin")
    b = b or Border()
    def keep(old, fallback):
        return old if (old and getattr(old, "style", None)) else fallback
    return Border(
        left   = keep(b.left,   thin),
        right  = keep(b.right,  thin),
        top    = keep(b.top,    thin),
        bottom = keep(b.bottom, thin),
    )

# Вертикальные и горизонтальные разделители (устойчивы к merge)

def draw_vertical_divider(ws, col, row_start, row_end, left=True, style="thin"):
    """
    Рисует вертикальную линию от row_start до row_end,
    НАЗНАЧАЯ границу КАЖДОЙ строке (включая MergedCell), чтобы линия не рвалась.
    """
    side = Side(style=style)
    for r in range(row_start, row_end + 1):
        cell = ws.cell(row=r, column=col)
        b = cell.border or Border()
        if left:
            cell.border = Border(left=side, right=b.right, top=b.top, bottom=b.bottom)
        else:
            cell.border = Border(left=b.left, right=side, top=b.top, bottom=b.bottom)

def draw_horizontal_divider(ws, row, col_start, col_end, style="thick"):
    """
    Жирная горизонтальная линия на границе row -> row+1 по всем колонкам (col_end включительно).
    Красим bottom у row и top у row+1, чтобы линия не рвалась из-за merge.
    """
    side = Side(style=style)
    for c in range(col_start, col_end + 1):
        # нижняя граница у строки row
        cell = ws.cell(row=row, column=c)
        b = cell.border or Border()
        cell.border = Border(left=b.left, right=b.right, top=b.top, bottom=side)
        # верхняя граница у строки row+1
        cell2 = ws.cell(row=row + 1, column=c)
        b2 = cell2.border or Border()
        cell2.border = Border(left=b2.left, right=b2.right, top=side, bottom=b2.bottom)

# Генератор отчётов

class ReportGenerator:
    def __init__(self, file_path):
        self.df_selection = read_excel_flexible(
            file_path,
            sheet_name='Sheet1'
        )

        # фильтрация колонок — ПОСЛЕ чтения
        self.df_selection = self.df_selection[
            [c for c in self.df_selection.columns if c not in ('Ссылка', 'Примечание')]
        ]

        # Поиск колонки с площадью/объёмом (га) — устойчиво к NBSP/латинице и т.д.
        self.area_col = None
        for col in self.df_selection.columns:
            col_norm = normalize_text(col)
            if ("га" in col_norm) and any(
                    x in col_norm for x in ("объём", "объем", "площадь", "сумма объем", "сумма объём")):
                # сохраняем оригинальное имя колонки (как в df), чтобы потом df[self.area_col] работал
                self.area_col = str(col).strip()
                break

        if not self.area_col:
            raise KeyError(f"❌ Не найден столбец с объёмом (га). Колонки: {list(self.df_selection.columns)}")

        # Поиск колонки с номером договора
        self.contract_col = None
        for col in self.df_selection.columns:
            s = normalize_text(col)
            if any(x in s for x in ['ген. договор', 'ген договор', 'номер дог', '№ ген. договора', 'ген.договор']):
                self.contract_col = str(col).strip()
                break

        if not self.contract_col:
            raise KeyError(f"❌ Не найден столбец 'Ген. договор'. Колонки: {list(self.df_selection.columns)}")

        # Тип ОГХ
        self.type_col = None
        for col in self.df_selection.columns:
            if 'тип объекта огх' in normalize_text(col):
                self.type_col = str(col).strip()
                break
        if not self.type_col:
            raise KeyError("❌ Не найден столбец 'Тип объекта ОГХ'.")

        # Отдел-исполнитель ИТП
        self.itp_col = None
        for col in self.df_selection.columns:
            s = str(col).lower()
            if 'отдел' in s and 'исполнитель' in s and 'итп' in s:
                self.itp_col = str(col).strip()
                break
        if not self.itp_col:
            self.itp_col = "Отдел исполнитель"
            self.df_selection[self.itp_col] = "Основной"

        # Фильтр входных
        order_col = '№ Заказа МГГТ'
        status_col = 'Состояние (действующий / приостановлен / аннулирован)'

        if order_col in self.df_selection.columns:
            has_R_or_K = self.df_selection[order_col].astype(str).str.contains(r'[РКВ]', na=False, case=False)
        else:
            has_R_or_K = pd.Series(False, index=self.df_selection.index)

        if status_col in self.df_selection.columns:
            is_annulled = self.df_selection[status_col].astype(str).str.lower().str.contains('аннулирован')
        else:
            is_annulled = pd.Series(False, index=self.df_selection.index)

        mask_keep = (~has_R_or_K) & (~is_annulled)
        self.selection_filtered = self.df_selection[mask_keep].copy()

        # Нормализуем тип ОГХ
        SELF_TYPE_NORM_COL = "_TYPE_NORM"
        self.selection_filtered[SELF_TYPE_NORM_COL] = (
            self.selection_filtered[self.type_col]
            .astype(str)
            .str.replace("\u00A0", " ", regex=False)  # NBSP -> пробел
            .str.strip()
            .str.replace(r"\s+", "", regex=True)      # убрать все пробелы
            .str.upper()
        )
        self.type_norm_col = SELF_TYPE_NORM_COL

        # Операции -> дата-колонка
        self.operation_map = {
            '1. Выдача нарядов': 'Дата наряда',
            '2. Выполнение полевых работ': 'Дата изготовления полевых работ по факту',
            '3. Изготовление ИТП': 'Дата операции Исправление после корректуры',
            '4. Определ. площадных и кол. характеристик (заливка)': 'Дата изготовления геоподосновы по факту',
            '5. Загрузка в САПР': 'Дата загрузки в САПР МГГТ',
            '6. Согласование границ в САПР': 'Дата согласования границ',
            '7. Загрузка в АСУ ОДС': 'Дата загрузки в АСУ ОДС (АСД)',
            '8. Утверждение в АСУ ОДС': 'Дата утверждения в АСУ ОДС (МГГТ)'
        }

        for col in self.operation_map.values():
            if col in self.selection_filtered.columns:
                self.selection_filtered[col] = pd.to_datetime(
                    self.selection_filtered[col],
                    errors='coerce',
                    format='%d.%m.%y'
                )

        # Группы операций
        self.operation_groups = {
            "ОДиПД": ["1. Выдача нарядов"],
            "УГП": ["2. Выполнение полевых работ", "3. Изготовление ИТП"],
            "ОП": [
                "4. Определ. площадных и кол. характеристик (заливка)",
                "5. Загрузка в САПР",
                "6. Согласование границ в САПР",
                "7. Загрузка в АСУ ОДС",
                "8. Утверждение в АСУ ОДС"
            ]
        }

        self.ops_with_itp_split = {"1. Выдача нарядов", "2. Выполнение полевых работ", "3. Изготовление ИТП"}

        # Стили и цвета
        self.group_colors = {"ОДиПД": "EBF1DE", "УГП": "DCE6F1", "ОП": "FCD5B4"}
        self.operation_fills = {}
        for g, ops in self.operation_groups.items():
            color = self.group_colors.get(g)
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for op in ops:
                self.operation_fills[op] = fill

        self.fact_fill = PatternFill(start_color="C6E0B4", fill_type="solid")
        self.percent_fill = PatternFill(start_color="B4C6E7", fill_type="solid")
        self.work_fill = PatternFill(start_color="FFFF99", fill_type="solid")
        self.subgroup_fill = PatternFill(start_color="FFE6CC", fill_type="solid")

        self.left = Alignment(horizontal="left",  vertical="center", wrap_text=True)
        self.center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.center_nowrap = Alignment(horizontal="center", vertical="center", wrap_text=False)  # для дат!
        self.bold = Font(bold=True)
        self.normal = Font(bold=False)
        self.thin = HeaderBuilder.thin

        self.remaining_fill = PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid")
        self.remaining_font = Font(color="9C0006", bold=True)

    def _pretty_itp(self, itp_val: str) -> str:
        s = str(itp_val or "")
        return "ОП" if "паспортизац" in s.lower() else s

    def _op_split_itp(self, op: str, split_by_itp: bool) -> bool:
        """
        Разбивку по ИТП включаем ТОЛЬКО для операций 1–3.
        """
        return bool(split_by_itp and (op in self.ops_with_itp_split))

    def filter_by_gz_year(self, df, gz_year):
        patterns = [
            f"12/ОГХ-{gz_year}",
            f"12/ОГХ-{2000 + gz_year}",
            f"ОГХ-{gz_year}",
            f"ОГХ-{2000 + gz_year}",
        ]
        col_data = df[self.contract_col].astype(str).str.lower()
        mask = pd.Series(False, index=df.index)
        for p in patterns:
            mask |= col_data.str.contains(p.lower(), na=False)
        return df[mask].copy()

    # шапка
    def _build_header(self, ws, periods, include_itp=False, include_ogx=False, daily=False):
        top = 1
        ws.row_dimensions[top].height = 40
        ws.row_dimensions[top + 1].height = 22

        HeaderBuilder.set_merged(ws, top, 1, top + 1, 1,
                                 "Ответственное подразделение", font=self.bold, border=self.thin, align=self.center)
        HeaderBuilder.set_merged(ws, top, 2, top + 1, 2,
                                 "Стадия выполнения/Период", font=self.bold, border=self.thin, align=self.center)

        current_col = 3
        if include_itp:
            HeaderBuilder.set_merged(ws, top, current_col, top + 1, current_col,
                                     "Отдел исполнитель", font=self.bold, border=self.thin, align=self.center)
            current_col += 1
        if include_ogx:
            HeaderBuilder.set_merged(ws, top, current_col, top + 1, current_col,
                                     "ТИП ОГХ", font=self.bold, border=self.thin, align=self.center)
            current_col += 1

        first_period_col = current_col
        # подписи периодов — НЕ объединяем, только строка 1 + "подложка" во 2-й строке
        for i, p in enumerate(periods):
            c = first_period_col + i
            cell = ws.cell(row=top, column=c)
            cell.value = p
            cell.font = self.normal
            cell.alignment = self.center_nowrap
            cell.border = self.thin
            # второй ряд шапки с тонкой рамкой
            cell2 = ws.cell(row=top + 1, column=c)
            cell2.border = self.thin
            cell2.alignment = self.center_nowrap

        fact_col = first_period_col + len(periods)
        percent_col = fact_col + 1
        work_col = percent_col + 1
        remaining_col = work_col + 1

        HeaderBuilder.set_merged(ws, top, fact_col, top + 1, fact_col,
                                 "Факт выполнения Графика", font=self.bold, fill=self.fact_fill, border=self.thin, align=self.center)
        HeaderBuilder.set_merged(ws, top, percent_col, top + 1, percent_col,
                                 "% выполнения графика", font=self.bold, fill=self.percent_fill, border=self.thin, align=self.center)
        HeaderBuilder.set_merged(ws, top, work_col, top + 1, work_col,
                                 "В работе по стадиям", font=self.bold, fill=self.work_fill, border=self.thin, align=self.center)
        HeaderBuilder.set_merged(ws, top, remaining_col, top + 1, remaining_col,
                                 "Осталось выполнить", font=self.bold, fill=self.remaining_fill, border=self.thin, align=self.center)

        return top + 2, first_period_col, fact_col, percent_col, work_col, remaining_col

    # отображаемые строки
    def _visible_rows_for_operation(self, rows_full, op, split_by_itp, split_by_ogx):
        """
        Управляет тем, какие строки реально показываем.
        В op1 при split_by_itp=True — показываем только «ВСЕГО».
        Для остальных операций — показываем всё, что сгенерировал _iter_rows_for_operation.
        """
        # по умолчанию показываем всё
        rows = list(rows_full)

        # 1-я операция: при разбивке по ИТП оставить только строку «ВСЕГО»
        if op == "1. Выдача нарядов" and split_by_itp:
            rows = [
                item for item in rows_full
                if isinstance(item, (list, tuple))
                   and len(item) == 2
                   and isinstance(item[1], (list, tuple))
                   and len(item[1]) >= 1
                   and item[1][0] == "ВСЕГО"
            ]

        return rows

    def _prev_total_for_labels(self, prev_map, labels):
        if not labels:
            return 0  #  защита от пустых меток
        key = tuple(labels)
        if key in prev_map:
            return prev_map[key]
        if len(labels) >= 1:
            last = labels[-1]
            total = 0
            for k, v in prev_map.items():
                if isinstance(k, tuple) and len(k) >= 1 and k[-1] == last:
                    total += v
            if total != 0:
                return total
        return 0

    def _iter_rows_for_operation(self, df, op, split_by_ogx, split_by_itp):
        #  подготовка списков ИТП и типов ОГХ
        # ИТП берём из исходного столбца, но приводим отображаемую метку через _pretty_itp
        raw_itp = df[self.itp_col].fillna("").astype(str)
        itp_deps = sorted(set(self._pretty_itp(x) for x in raw_itp if x.strip()))

        # Типы ОГХ берём из нормализованного столбца (как у тебя уже сделано)
        type_series = df[self.type_norm_col].dropna().astype(str)
        dyn_types = sorted([x for x in type_series.unique() if x], key=str)
        preferred = ["ДТ", "ОО", "ОДХ"]
        ogx_types = preferred + [t for t in dyn_types if t not in preferred]

        #  маски
        def mask_total():
            return pd.Series(True, index=df.index)

        def make_mask_by_type(type_val):
            tv = str(type_val).replace("\u00A0", " ").strip()
            tv = "".join(tv.split()).upper()
            return lambda: (df[self.type_norm_col] == tv)

        def make_mask_by_itp(itp_val):
            # сравнение по отрисовываемому значению ИТП (через _pretty_itp)
            return lambda: (raw_itp.apply(self._pretty_itp) == itp_val)

        def make_mask_by_itp_and_type(itp_val, type_val):
            tv = str(type_val).replace("\u00A0", " ").strip()
            tv = "".join(tv.split()).upper()
            return lambda: (raw_itp.apply(self._pretty_itp) == itp_val) & (df[self.type_norm_col] == tv)

        show_itp = split_by_itp and (op in self.ops_with_itp_split)

        # Операция 1: ВИДИМ как раньше (ВСЕГО и, опционально, типы),
        # но ДОБАВЛЯЕМ "скрытые" строки ИТП (и ИТП×тип) в rows_full для расчётов «В работе»
        if op == "1. Выдача нарядов":
            # видимая «ВСЕГО»
            yield mask_total, ["ВСЕГО"]
            # видимые типы (если включена разбивка по ОГХ)
            if split_by_ogx:
                for t in ogx_types:
                    yield make_mask_by_type(t), [t]

            # скрытые ИТП — только для расчётов (будут отфильтрованы в _visible_rows_for_operation)
            if show_itp:
                for itp in itp_deps:
                    yield make_mask_by_itp(itp), [itp]
                    if split_by_ogx:
                        for t in ogx_types:
                            yield make_mask_by_itp_and_type(itp, t), [itp, t]
            return

        #  Операции 2+ (логика как была, только ИТП теперь включает и «ОП»)
        if not split_by_ogx and not split_by_itp:
            yield mask_total, ["ВСЕГО"]

        elif not split_by_ogx and split_by_itp:
            yield mask_total, ["ВСЕГО"]
            for itp in itp_deps:
                yield make_mask_by_itp(itp), [itp]

        elif split_by_ogx and not split_by_itp:
            yield mask_total, ["ВСЕГО"]
            for t in ogx_types:
                yield make_mask_by_type(t), [t]

        else:
            yield mask_total, ["ВСЕГО"]
            for itp in itp_deps:
                yield make_mask_by_itp(itp), [itp]
                for t in ogx_types:
                    yield make_mask_by_itp_and_type(itp, t), [itp, t]

    # Безопасный итератор по строкам вида (mask_func, labels)
    def _safe_iter(self, items):
        for item in items:
            if isinstance(item, (list, tuple)) and len(item) == 2:
                yield item

    def _auto_width(self, ws):
        # только ширины
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) for cell in col
                 if cell.value is not None and not isinstance(cell, MergedCell)),
                default=0
            )
            width = max(10, min(max_len + 2, 60))
            ws.column_dimensions[get_column_letter(col[0].column)].width = width

        # Сделаем A и B пошире
        ws.column_dimensions['A'].width = max(ws.column_dimensions['A'].width or 0, 28)
        ws.column_dimensions['B'].width = max(ws.column_dimensions['B'].width or 0, 56)

        #  аккуратно «дорисуем» тонкую сетку, сохранив thick
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                cell.border = _merge_with_thin_preserving(cell.border)

    def _unify_row_heights(self, ws, start_row=3, default_h=22):
        for r in range(start_row, ws.max_row + 1):
            ws.row_dimensions[r].height = default_h

    def _hide_all_but_last_n_periods(self, ws, first_period_col, periods_count, keep_last=4):
        hide_until = max(0, periods_count - keep_last)
        for c in range(first_period_col, first_period_col + hide_until):
            ws.column_dimensions[get_column_letter(c)].hidden = True

    # сводные %/Осталось при разбивке
    def _merge_summary_cols_for_block(self, ws, row_start, row_end, col_percent, col_remain, percent_value, remain_value):

        if row_end <= row_start:
            return
        HeaderBuilder.set_merged(ws, row_start, col_percent, row_end, col_percent,
                                 f"{percent_value}%", font=self.bold, fill=self.percent_fill,
                                 align=self.center, border=self.thin)
        HeaderBuilder.set_merged(ws, row_start, col_remain, row_end, col_remain,
                                 remain_value, font=self.remaining_font, fill=self.remaining_fill,
                                 align=self.center, border=self.thin)

    #  ЕЖЕДНЕВНЫЙ
    def generate_daily(self, gz_year, plan_count, plan_area, split_by_ogx=False, split_by_itp=False):
        df = self.filter_by_gz_year(self.selection_filtered, gz_year)
        today = datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)
        weekday = today.weekday()
        end_date = today - timedelta(days=1) if weekday == 0 else (today - timedelta(days=weekday) + timedelta(days=6))
        start_date = end_date - timedelta(days=20)
        date_range = pd.date_range(start=start_date, end=end_date, freq='D')

        months_ru = ["янв", "фев", "мар", "апр", "май", "июн", "июл", "авг", "сен", "окт", "ноя", "дек"]
        periods = [f"{d.day} {months_ru[d.month-1]}" for d in date_range]

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = f"ежедневный_отчет_{gz_year}_{timestamp}.xlsx"

        with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
            wb = writer.book
            ws_area = wb.create_sheet('Га.')
            ws_count = wb.create_sheet('Шт.')

            row_a, base_a, fact_a, percent_a, work_a, remain_a = self._build_header(ws_area, periods, include_itp=split_by_itp, include_ogx=split_by_ogx, daily=True)
            row_c, base_c, fact_c, percent_c, work_c, remain_c = self._build_header(ws_count, periods, include_itp=split_by_itp, include_ogx=split_by_ogx, daily=True)
            header_bottom_a = row_a - 1
            header_bottom_c = row_c - 1

            row_area = row_a
            row_count = row_c

            prev_fact_count_by_group = {}
            prev_fact_area_by_group = {}

            for group_name, ops in self.operation_groups.items():
                group_start_row_area = row_area
                group_start_row_count = row_count

                for op in ops:
                    if op not in self.operation_map or self.operation_map[op] not in df.columns:
                        continue

                    for c in range(1, remain_a + 1):
                        _set_top_thick(ws_area, row_area, c)
                    for c in range(1, remain_c + 1):
                        _set_top_thick(ws_count, row_count, c)

                    # ЖИРНАЯ линия в НАЧАЛЕ операции (между row-1 и row)
                    draw_horizontal_divider(ws_area, row_area - 1, 1, remain_a, style="thick")
                    draw_horizontal_divider(ws_count, row_count - 1, 1, remain_c, style="thick")

                    col_name = self.operation_map[op]
                    fill = self.operation_fills[op]

                    # локальный флаг: разбивка по ИТП только для 1–3 операций
                    op_split_itp = self._op_split_itp(op, split_by_itp)
                    show_itp_col = op_split_itp

                    rows_full = list(self._iter_rows_for_operation(df, op, split_by_ogx, op_split_itp))

                    current_fact_count_by_group = {}
                    current_fact_area_by_group  = {}
                    for mask_func, labels in self._safe_iter(rows_full):
                        mask = mask_func()
                        fact_mask = mask & df[col_name].notna()
                        fact_count = int(fact_mask.sum())
                        fact_area  = truncate_float(float(df.loc[fact_mask, self.area_col].fillna(0).sum()), 1)
                        current_fact_count_by_group[tuple(labels)] = fact_count
                        current_fact_area_by_group[tuple(labels)]  = fact_area

                    rows = list(self._visible_rows_for_operation(rows_full, op, op_split_itp, split_by_ogx))
                    rows_safe = list(self._safe_iter(rows))

                    row_area_temp = row_area
                    row_count_temp = row_count

                    prev_itp = None
                    total_percent_c = total_percent_a = 0
                    total_remain_c = total_remain_a = 0


                    for mask_func, labels in rows_safe:
                        mask = mask_func()
                        is_total = (len(labels) > 0 and labels[0] == "ВСЕГО")
                        write_col = 3
                        group_fill = PatternFill(start_color=self.group_colors.get(group_name, "FFFFFF"),
                                                 end_color=self.group_colors.get(group_name, "FFFFFF"),
                                                 fill_type="solid")

                        def style_cell(ws, r, c, val, bold=False, filled=False, align_left=False):
                            cell = HeaderBuilder.set_cell_value_safe(ws, r, c, val)
                            cell.alignment = self.left if align_left else self.center
                            cell.font = self.bold if bold else self.normal
                            cell.border = self.thin
                            if filled:
                                cell.fill = self.subgroup_fill
                            return cell

                        double_split = (split_by_ogx and show_itp_col)

                        if double_split:
                            if len(labels) == 1 and labels[0] == "ВСЕГО":
                                style_cell(ws_count, row_count_temp, write_col + 1, "ВСЕГО", bold=True, filled=True, align_left=True)
                                style_cell(ws_area, row_area_temp, write_col + 1, "ВСЕГО", bold=True, filled=True, align_left=True)
                            elif len(labels) == 1:
                                if op == "1. Выдача нарядов":
                                    # для 1-й операции это не ИТП, а тип ОГХ → смещаем на колонку вправо,
                                    # если в шапке есть "Отдел исполнитель"
                                    ogx_col = write_col + 1 if split_by_itp else write_col
                                    style_cell(ws_count, row_count_temp, ogx_col, labels[0], align_left=True)
                                    style_cell(ws_area, row_area_temp, ogx_col, labels[0], align_left=True)
                                    # важное: считаем эту колонку занятой
                                    write_col = max(write_col, ogx_col) + 1
                                elif len(labels) == 2:
                                    itp, ogx = labels
                                    if itp != prev_itp:
                                        # ваши вызовы style_cell/put для ITP
                                        style_cell(ws_count, row_count_temp, write_col, itp, bold=True, filled=True,
                                                   align_left=True)
                                        style_cell(ws_area, row_area_temp, write_col, itp, bold=True, filled=True,
                                                   align_left=True)
                                        prev_itp = itp
                                    # ваши вызовы style_cell/put для OGX
                                    style_cell(ws_count, row_count_temp, write_col + 1, ogx, align_left=True)
                                    style_cell(ws_area, row_area_temp, write_col + 1, ogx, align_left=True)
                                else:
                                    # запасной вариант
                                    style_cell(ws_count, row_count_temp, write_col, "ВСЕГО", bold=True, filled=True,
                                               align_left=True)
                                    style_cell(ws_area, row_area_temp, write_col, "ВСЕГО", bold=True, filled=True,
                                               align_left=True)


                            else:
                                itp, ogx = labels
                                if itp != prev_itp:
                                    style_cell(ws_count, row_count_temp, write_col, itp, bold=True, filled=True, align_left=True)
                                    style_cell(ws_area, row_area_temp, write_col, itp, bold=True, filled=True, align_left=True)
                                    prev_itp = itp
                                style_cell(ws_count, row_count_temp, write_col + 1, ogx, align_left=True)
                                style_cell(ws_area, row_area_temp, write_col + 1, ogx, align_left=True)
                        else:
                            # корректная колонка для ОГХ
                            if split_by_ogx and split_by_itp and (op not in self.ops_with_itp_split):
                                ogx_col = write_col + 1  # пропускаем колонку "Отдел исполнитель"
                            else:
                                ogx_col = write_col

                            if show_itp_col and len(labels) == 1 and labels[0] != "ВСЕГО":
                                lbl = labels[0]
                                # FIX: правильный расчёт количества подстрок ИТП
                                itp_count = sum(1 for _, labs in rows_safe if (len(labs) > 0 and labs[0] == lbl))
                                end_row = row_count_temp + itp_count - 1
                                HeaderBuilder.set_merged(ws_count, row_count_temp, write_col, end_row, write_col, lbl,
                                                         font=self.bold, fill=self.subgroup_fill, align=self.center, border=self.thin)
                                HeaderBuilder.set_merged(ws_area, row_area_temp, write_col, end_row, write_col, lbl,
                                                         font=self.bold, fill=self.subgroup_fill, align=self.center, border=self.thin)
                                write_col += 1

                            if split_by_ogx:
                                lbl = labels[-1] if len(labels) > 1 else labels[0]
                                HeaderBuilder.set_merged(ws_count, row_count_temp, ogx_col, row_count_temp, ogx_col, lbl,
                                                         font=self.bold if lbl == "ВСЕГО" else self.normal,
                                                         fill=self.subgroup_fill if lbl == "ВСЕГО" else None,
                                                         align=self.center, border=self.thin)
                                HeaderBuilder.set_merged(ws_area, row_area_temp, ogx_col, row_area_temp, ogx_col, lbl,
                                                         font=self.bold if lbl == "ВСЕГО" else self.normal,
                                                         fill=self.subgroup_fill if lbl == "ВСЕГО" else None,
                                                         align=self.center, border=self.thin)
                                # колонка ОГХ считается «занятой»
                                write_col = max(write_col, ogx_col) + 1
                            else:
                                # даже если ОГХ не разбиваем, подпись строки (например, "ВСЕГО") должна быть
                                lbl = labels[-1] if len(labels) > 1 else (labels[0] if labels else "")
                                HeaderBuilder.set_merged(ws_count, row_count_temp, ogx_col, row_count_temp, ogx_col,
                                                         lbl,
                                                         font=self.bold if lbl == "ВСЕГО" else self.normal,
                                                         fill=self.subgroup_fill if lbl == "ВСЕГО" else None,
                                                         align=self.center, border=self.thin)
                                HeaderBuilder.set_merged(ws_area, row_area_temp, ogx_col, row_area_temp, ogx_col, lbl,
                                                         font=self.bold if lbl == "ВСЕГО" else self.normal,
                                                         fill=self.subgroup_fill if lbl == "ВСЕГО" else None,
                                                         align=self.center, border=self.thin)
                                write_col = max(write_col, ogx_col) + 1

                        for j, date in enumerate(date_range):
                            date_mask = mask & (df[col_name].dt.date == date.date())
                            cnt = int(date_mask.sum())
                            area = float(df.loc[date_mask, self.area_col].fillna(0).sum())

                            cell_cnt = HeaderBuilder.set_cell_value_safe(ws_count, row_count_temp, base_c + j, cnt)
                            cell_cnt.alignment = self.center; cell_cnt.border = self.thin; cell_cnt.fill = fill
                            if is_total: cell_cnt.font = self.bold

                            cell_area = HeaderBuilder.set_cell_value_safe(ws_area, row_area_temp, base_a + j, truncate_float(area, 1))
                            cell_area.alignment = self.center; cell_area.border = self.thin; cell_area.fill = fill
                            if is_total: cell_area.font = self.bold

                        key = tuple(labels)
                        fact_count = current_fact_count_by_group.get(key, 0)
                        fact_area  = current_fact_area_by_group.get(key, 0.0)

                        if is_total:
                            percent_count = round((fact_count / plan_count * 100), 1) if plan_count > 0 else 0
                            percent_area  = round((fact_area  / plan_area  * 100), 1) if plan_area  > 0 else 0
                        else:
                            percent_count = round((fact_count / plan_count * 100), 2) if plan_count > 0 else 0
                            percent_area  = round((fact_area  / plan_area  * 100), 2) if plan_area  > 0 else 0

                        work_count = "-"
                        work_area  = "-"
                        if op != "1. Выдача нарядов":
                            prev_count = self._prev_total_for_labels(prev_fact_count_by_group, labels)
                            prev_area  = self._prev_total_for_labels(prev_fact_area_by_group,  labels)
                            work_count = max(prev_count - fact_count, 0)
                            work_area  = truncate_float(max(prev_area - fact_area, 0.0), 1)

                        remaining_count = max(plan_count - fact_count, 0)
                        remaining_area  = truncate_float(max(plan_area  - fact_area,  0.0), 1)

                        cell_fact_c = HeaderBuilder.set_cell_value_safe(ws_count, row_count_temp, fact_c, fact_count)
                        cell_fact_c.font = self.bold; cell_fact_c.fill = self.fact_fill; cell_fact_c.alignment = self.center; cell_fact_c.border = self.thin

                        cell_pc_c = HeaderBuilder.set_cell_value_safe(ws_count, row_count_temp, percent_c, f"{percent_count}%")
                        cell_pc_c.font = self.bold; cell_pc_c.fill = self.percent_fill; cell_pc_c.alignment = self.center; cell_pc_c.border = self.thin

                        cell_work_c = HeaderBuilder.set_cell_value_safe(ws_count, row_count_temp, work_c, work_count)
                        cell_work_c.font = self.bold; cell_work_c.fill = self.work_fill; cell_work_c.alignment = self.center; cell_work_c.border = self.thin

                        cell_rem_c = HeaderBuilder.set_cell_value_safe(ws_count, row_count_temp, remain_c, remaining_count)
                        cell_rem_c.font = self.remaining_font; cell_rem_c.fill = self.remaining_fill; cell_rem_c.alignment = self.center; cell_rem_c.border = self.thin

                        cell_fact_a = HeaderBuilder.set_cell_value_safe(ws_area, row_area_temp, fact_a, fact_area)
                        cell_fact_a.font = self.bold; cell_fact_a.fill = self.fact_fill; cell_fact_a.alignment = self.center; cell_fact_a.border = self.thin

                        cell_pc_a = HeaderBuilder.set_cell_value_safe(ws_area, row_area_temp, percent_a, f"{percent_area}%")
                        cell_pc_a.font = self.bold; cell_pc_a.fill = self.percent_fill; cell_pc_a.alignment = self.center; cell_pc_a.border = self.thin

                        cell_work_a = HeaderBuilder.set_cell_value_safe(ws_area, row_area_temp, work_a, work_area)
                        cell_work_a.font = self.bold; cell_work_a.fill = self.work_fill; cell_work_a.alignment = self.center; cell_work_a.border = self.thin

                        cell_rem_a = HeaderBuilder.set_cell_value_safe(ws_area, row_area_temp, remain_a, remaining_area)
                        cell_rem_a.font = self.remaining_font; cell_rem_a.fill = self.remaining_fill; cell_rem_a.alignment = self.center; cell_rem_a.border = self.thin

                        if is_total:
                            total_percent_c = percent_count
                            total_percent_a = percent_area
                            total_remain_c = remaining_count
                            total_remain_a = remaining_area

                        HeaderBuilder._apply_range(ws_count, row_count_temp, 1, row_count_temp, fact_c - 1, fill=group_fill)
                        HeaderBuilder._apply_range(ws_area,  row_area_temp,  1, row_area_temp,  fact_a - 1, fill=group_fill)

                        row_count_temp += 1
                        row_area_temp  += 1

                    op_end_row_area  = row_area_temp  - 1
                    op_end_row_count = row_count_temp - 1

                    HeaderBuilder.set_merged(ws_area,  row_area, 2, op_end_row_area,  2, op, font=self.bold, fill=self.operation_fills[op], align=self.left,  border=self.thin)
                    HeaderBuilder.set_merged(ws_count, row_count, 2, op_end_row_count, 2, op, font=self.bold, fill=self.operation_fills[op], align=self.left,  border=self.thin)

                    has_split = (split_by_itp or split_by_ogx)
                    if has_split and op_end_row_count > row_count:
                        self._merge_summary_cols_for_block(ws_count, row_count, op_end_row_count, percent_c, remain_c, total_percent_c, total_remain_c)
                        self._merge_summary_cols_for_block(ws_area,  row_area,  op_end_row_area,  percent_a, remain_a, total_percent_a, total_remain_a)

                    # --- ЖИРНАЯ горизонтальная линия в конце операции (без разрывов)
                    draw_horizontal_divider(ws_area,  op_end_row_area, 1, remain_a, style="thick")
                    draw_horizontal_divider(ws_count, op_end_row_count, 1, remain_c, style="thick")

                    row_area  = row_area_temp + 1
                    row_count = row_count_temp + 1

                    prev_fact_count_by_group = current_fact_count_by_group
                    prev_fact_area_by_group  = current_fact_area_by_group

                if row_area - group_start_row_area > 1:
                    fill = PatternFill(start_color=self.group_colors.get(group_name, "FFFFFF"),
                                       end_color=self.group_colors.get(group_name, "FFFFFF"),
                                       fill_type="solid")
                    HeaderBuilder.set_merged(ws_area, group_start_row_area, 1, row_area - 1, 1, group_name,
                                             font=self.bold, align=self.center, border=self.thin, fill=fill)
                    HeaderBuilder.set_merged(ws_count, group_start_row_count, 1, row_count - 1, 1, group_name,
                                             font=self.bold, align=self.center, border=self.thin, fill=fill)

            for ws in [ws_area, ws_count]:
                self._auto_width(ws)
                self._unify_row_heights(ws, start_row=3, default_h=22)

            # тонкая линия строго под шапкой — по всей ширине
            draw_horizontal_divider(ws_area,  header_bottom_a, 1, remain_a, style="thick")
            draw_horizontal_divider(ws_count, header_bottom_c, 1, remain_c, style="thick")

            #  вертикальные разделители

            # между A и B (тонкая)
            draw_vertical_divider(ws_area,  2, 1, ws_area.max_row, left=True, style="thin")
            draw_vertical_divider(ws_count, 2, 1, ws_count.max_row, left=True, style="thin")

            # между последней датой и "Факт" (тонкая)
            draw_vertical_divider(ws_area,  base_a, 1, ws_area.max_row, left=True, style="thin")
            draw_vertical_divider(ws_count, base_c, 1, ws_count.max_row, left=True, style="thin")

            # внутри финальных колонок (тонкие)
            for col in (fact_a, percent_a, work_a, remain_a):
                draw_vertical_divider(ws_area,  col, 1, ws_area.max_row, left=True, style="thin")
            for col in (fact_c, percent_c, work_c, remain_c):
                draw_vertical_divider(ws_count, col, 1, ws_count.max_row, left=True, style="thin")

            # понедельничные разделители (тонкие) — опционально
            for j, d in enumerate(date_range):
                if d.weekday() == 0 and j > 0:
                    draw_vertical_divider(ws_area,  base_a + j, 1, ws_area.max_row, left=True, style="thin")
                    draw_vertical_divider(ws_count, base_c + j, 1, ws_count.max_row, left=True, style="thin")

            #  ЖИРНАЯ вертикальная линия в конце каждой недели (после воскресенья)
            for j, d in enumerate(date_range):
                if d.weekday() == 6:  # воскресенье
                    col_after_a = (base_a + j + 1) if (j + 1) < len(date_range) else fact_a
                    col_after_c = (base_c + j + 1) if (j + 1) < len(date_range) else fact_c
                    draw_vertical_divider(ws_area,  col_after_a, 1, ws_area.max_row, left=True, style="thick")
                    draw_vertical_divider(ws_count, col_after_c, 1, ws_count.max_row, left=True, style="thick")

            # внешняя толстая рамка
            draw_outline(ws_area, 1, 1, ws_area.max_row, remain_a)
            draw_outline(ws_count, 1, 1, ws_count.max_row, remain_c)

            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']

        open_file_crossplatform(out_path)

    # НЕДЕЛЬНЫЙ
    def generate_weekly_combined(self, gz_year, plan_count, plan_area, split_by_ogx=False, split_by_itp=False):
        df = self.filter_by_gz_year(self.selection_filtered, gz_year)
        year_full = 2000 + int(gz_year)
        start_date = datetime(year_full, 1, 1)
        if "Дата наряда" in df.columns:
            mask_dec = (df["Дата наряда"].dt.year == year_full - 1) & (df["Дата наряда"].dt.month == 12)
            if mask_dec.any():
                start_date = datetime(year_full - 1, 12, 2)
        today = datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)
        weeks = pd.date_range(start=start_date, end=today, freq='W-MON')
        periods = [f"{w.strftime('%d.%m.%Y')}-{(w + timedelta(days=6)).strftime('%d.%m.%Y')}" for w in weeks]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = f"недельный_отчет_{gz_year}_{timestamp}.xlsx"

        with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
            wb = writer.book
            ws_area = wb.create_sheet('Га.')
            ws_count = wb.create_sheet('Шт.')

            row_a, base_a, fact_a, percent_a, work_a, remain_a = self._build_header(ws_area, periods, include_itp=split_by_itp, include_ogx=split_by_ogx)
            row_c, base_c, fact_c, percent_c, work_c, remain_c = self._build_header(ws_count, periods, include_itp=split_by_itp, include_ogx=split_by_ogx)
            header_bottom_a = row_a - 1
            header_bottom_c = row_c - 1

            row_area = row_a
            row_count = row_c

            prev_fact_count_by_group = {}
            prev_fact_area_by_group = {}

            for group_name, ops in self.operation_groups.items():
                group_start_row_area = row_area
                group_start_row_count = row_count

                for op in ops:
                    if op not in self.operation_map or self.operation_map[op] not in df.columns:
                        continue

                    for c in range(1, remain_a + 1):
                        _set_top_thick(ws_area, row_area, c)
                    for c in range(1, remain_c + 1):
                        _set_top_thick(ws_count, row_count, c)

                    # ЖИРНАЯ линия в НАЧАЛЕ операции (между row-1 и row)
                    draw_horizontal_divider(ws_area, row_area - 1, 1, remain_a, style="thick")
                    draw_horizontal_divider(ws_count, row_count - 1, 1, remain_c, style="thick")

                    col_name = self.operation_map[op]
                    fill = self.operation_fills[op]

                    # локальный флаг: разбивка по ИТП только для 1–3 операций
                    op_split_itp = self._op_split_itp(op, split_by_itp)
                    show_itp_col = op_split_itp

                    rows_full = list(self._iter_rows_for_operation(df, op, split_by_ogx, op_split_itp))


                    current_fact_count_by_group = {}
                    current_fact_area_by_group  = {}
                    for mask_func, labels in self._safe_iter(rows_full):
                        mask = mask_func()
                        fact_mask = mask & df[col_name].notna()
                        fact_count = int(fact_mask.sum())
                        fact_area  = truncate_float(float(df.loc[fact_mask, self.area_col].fillna(0).sum()), 1)
                        current_fact_count_by_group[tuple(labels)] = fact_count
                        current_fact_area_by_group[tuple(labels)]  = fact_area

                    rows = list(self._visible_rows_for_operation(rows_full, op, op_split_itp, split_by_ogx))
                    rows_safe = list(self._safe_iter(rows))

                    row_area_temp = row_area
                    row_count_temp = row_count
                    prev_itp = None

                    total_percent_c = total_percent_a = 0
                    total_remain_c = total_remain_a = 0

                    for mask_func, labels in self._safe_iter(rows):
                        mask = mask_func()
                        is_total = (len(labels) > 0 and labels[0] == "ВСЕГО")
                        write_col = 3
                        group_fill = PatternFill(start_color=self.group_colors.get(group_name, "FFFFFF"),
                                                 end_color=self.group_colors.get(group_name, "FFFFFF"),
                                                 fill_type="solid")

                        def put(ws, r, c, val, bold=False, filled=False, align_left=False):
                            cell = HeaderBuilder.set_cell_value_safe(ws, r, c, val)
                            cell.alignment = self.left if align_left else self.center
                            cell.font = self.bold if bold else self.normal
                            cell.border = self.thin
                            if filled:
                                cell.fill = self.subgroup_fill
                            return cell

                        double_split = (split_by_ogx and show_itp_col)

                        if double_split:
                            if len(labels) == 0:
                                pass  # нечего выводить, просто пропускаем строку
                            elif len(labels) == 1 and labels[0] == "ВСЕГО":
                                put(ws_count, row_count_temp, write_col + 1, "ВСЕГО", bold=True, filled=True,
                                    align_left=True)
                                put(ws_area, row_area_temp, write_col + 1, "ВСЕГО", bold=True, filled=True,
                                    align_left=True)
                            elif len(labels) == 1:
                                if op == "1. Выдача нарядов":
                                    ogx_col = write_col + 1 if split_by_itp else write_col
                                    put(ws_count, row_count_temp, ogx_col, labels[0], align_left=True)
                                    put(ws_area, row_area_temp, ogx_col, labels[0], align_left=True)
                                    write_col = max(write_col, ogx_col) + 1
                                else:
                                    itp = labels[0]
                                    put(ws_count, row_count_temp, write_col, itp, bold=True, filled=True,
                                        align_left=True)
                                    put(ws_area, row_area_temp, write_col, itp, bold=True, filled=True, align_left=True)
                                    prev_itp = itp
                            elif len(labels) == 2:
                                itp, ogx = labels
                                if itp != prev_itp:
                                    put(ws_count, row_count_temp, write_col, itp, bold=True, filled=True,
                                        align_left=True)
                                    put(ws_area, row_area_temp, write_col, itp, bold=True, filled=True, align_left=True)
                                    prev_itp = itp
                                put(ws_count, row_count_temp, write_col + 1, ogx, align_left=True)
                                put(ws_area, row_area_temp, write_col + 1, ogx, align_left=True)
                            else:
                                # Защита на случай неожиданной длины labels
                                put(ws_count, row_count_temp, write_col, "ВСЕГО", bold=True, filled=True,
                                    align_left=True)
                                put(ws_area, row_area_temp, write_col, "ВСЕГО", bold=True, filled=True, align_left=True)

                        else:
                            # корректная колонка для ОГХ
                            if split_by_ogx and split_by_itp and (op not in self.ops_with_itp_split):
                                ogx_col = write_col + 1
                            else:
                                ogx_col = write_col

                            if op_split_itp and len(labels) == 1 and labels[0] != "ВСЕГО":
                                lbl = labels[0]
                                itp_count = sum(1 for _, labs in rows_safe if (len(labs) > 0 and labs[0] == lbl))
                                end_row = row_count_temp + itp_count - 1
                                HeaderBuilder.set_merged(ws_count, row_count_temp, write_col, end_row, write_col, lbl,
                                                         font=self.bold, fill=self.subgroup_fill, align=self.center,
                                                         border=self.thin)
                                HeaderBuilder.set_merged(ws_area, row_area_temp, write_col, end_row, write_col, lbl,
                                                         font=self.bold, fill=self.subgroup_fill, align=self.center,
                                                         border=self.thin)
                                write_col += 1

                            if split_by_ogx:
                                lbl = labels[-1] if len(labels) > 1 else (labels[0] if labels else "")
                                HeaderBuilder.set_merged(ws_count, row_count_temp, ogx_col, row_count_temp, ogx_col, lbl,
                                                         font=self.bold if lbl == "ВСЕГО" else self.normal,
                                                         fill=self.subgroup_fill if lbl == "ВСЕГО" else None,
                                                         align=self.center, border=self.thin)
                                HeaderBuilder.set_merged(ws_area, row_area_temp, ogx_col, row_area_temp, ogx_col, lbl,
                                                         font=self.bold if lbl == "ВСЕГО" else self.normal,
                                                         fill=self.subgroup_fill if lbl == "ВСЕГО" else None,
                                                         align=self.center, border=self.thin)
                                write_col = max(write_col, ogx_col) + 1
                            else:
                                lbl = labels[-1] if len(labels) > 1 else (labels[0] if labels else "")
                                HeaderBuilder.set_merged(ws_count, row_count_temp, ogx_col, row_count_temp, ogx_col,
                                                         lbl,
                                                         font=self.bold if lbl == "ВСЕГО" else self.normal,
                                                         fill=self.subgroup_fill if lbl == "ВСЕГО" else None,
                                                         align=self.center, border=self.thin)
                                HeaderBuilder.set_merged(ws_area, row_area_temp, ogx_col, row_area_temp, ogx_col, lbl,
                                                         font=self.bold if lbl == "ВСЕГО" else self.normal,
                                                         fill=self.subgroup_fill if lbl == "ВСЕГО" else None,
                                                         align=self.center, border=self.thin)
                                write_col = max(write_col, ogx_col) + 1

                        for j, week_start in enumerate(weeks):
                            week_end = week_start + timedelta(days=6)
                            week_mask = mask & (df[col_name] >= week_start) & (df[col_name] <= week_end)
                            cnt = int(week_mask.sum())
                            area = float(df.loc[week_mask, self.area_col].fillna(0).sum())

                            cell_cnt = HeaderBuilder.set_cell_value_safe(ws_count, row_count_temp, base_c + j, cnt)
                            cell_cnt.alignment = self.center; cell_cnt.border = self.thin; cell_cnt.fill = fill
                            if is_total: cell_cnt.font = self.bold

                            cell_area = HeaderBuilder.set_cell_value_safe(ws_area, row_area_temp, base_a + j, truncate_float(area, 1))
                            cell_area.alignment = self.center; cell_area.border = self.thin; cell_area.fill = fill
                            if is_total: cell_area.font = self.bold

                        key = tuple(labels)
                        fact_count = current_fact_count_by_group.get(key, 0)
                        fact_area  = current_fact_area_by_group.get(key, 0.0)

                        if is_total:
                            percent_count = round((fact_count / plan_count * 100), 1) if plan_count > 0 else 0
                            percent_area  = round((fact_area  / plan_area  * 100), 1) if plan_area  > 0 else 0
                        else:
                            percent_count = round((fact_count / plan_count * 100), 2) if plan_count > 0 else 0
                            percent_area  = round((fact_area  / plan_area  * 100), 2) if plan_area  > 0 else 0

                        work_count = "-"
                        work_area  = "-"
                        if op != "1. Выдача нарядов":
                            prev_count = self._prev_total_for_labels(prev_fact_count_by_group, labels)
                            prev_area  = self._prev_total_for_labels(prev_fact_area_by_group,  labels)
                            work_count = max(prev_count - fact_count, 0)
                            work_area  = truncate_float(max(prev_area - fact_area, 0.0), 1)

                        remaining_count = max(plan_count - fact_count, 0)
                        remaining_area  = truncate_float(max(plan_area  - fact_area,  0.0), 1)

                        HeaderBuilder.set_cell_value_safe(ws_count, row_count_temp, fact_c, fact_count).font = self.bold
                        ws_count.cell(row=row_count_temp, column=fact_c).fill = self.fact_fill
                        ws_count.cell(row=row_count_temp, column=fact_c).alignment = self.center
                        ws_count.cell(row=row_count_temp, column=fact_c).border = self.thin

                        HeaderBuilder.set_cell_value_safe(ws_count, row_count_temp, percent_c, f"{percent_count}%").font = self.bold
                        ws_count.cell(row=row_count_temp, column=percent_c).fill = self.percent_fill
                        ws_count.cell(row=row_count_temp, column=percent_c).alignment = self.center
                        ws_count.cell(row=row_count_temp, column=percent_c).border = self.thin

                        HeaderBuilder.set_cell_value_safe(ws_count, row_count_temp, work_c, work_count).font = self.bold
                        ws_count.cell(row=row_count_temp, column=work_c).fill = self.work_fill
                        ws_count.cell(row=row_count_temp, column=work_c).alignment = self.center
                        ws_count.cell(row=row_count_temp, column=work_c).border = self.thin

                        HeaderBuilder.set_cell_value_safe(ws_count, row_count_temp, remain_c, remaining_count).font = self.remaining_font
                        ws_count.cell(row=row_count_temp, column=remain_c).fill = self.remaining_fill
                        ws_count.cell(row=row_count_temp, column=remain_c).alignment = self.center
                        ws_count.cell(row=row_count_temp, column=remain_c).border = self.thin

                        HeaderBuilder.set_cell_value_safe(ws_area, row_area_temp, fact_a, fact_area).font = self.bold
                        ws_area.cell(row=row_area_temp, column=fact_a).fill = self.fact_fill
                        ws_area.cell(row=row_area_temp, column=fact_a).alignment = self.center
                        ws_area.cell(row=row_area_temp, column=fact_a).border = self.thin

                        HeaderBuilder.set_cell_value_safe(ws_area, row=row_area_temp, col=percent_a, value=f"{percent_area}%").font = self.bold
                        ws_area.cell(row=row_area_temp, column=percent_a).fill = self.percent_fill
                        ws_area.cell(row=row_area_temp, column=percent_a).alignment = self.center
                        ws_area.cell(row=row_area_temp, column=percent_a).border = self.thin

                        HeaderBuilder.set_cell_value_safe(ws_area, row_area_temp, work_a, work_area).font = self.bold
                        ws_area.cell(row=row_area_temp, column=work_a).fill = self.work_fill
                        ws_area.cell(row=row_area_temp, column=work_a).alignment = self.center
                        ws_area.cell(row=row_area_temp, column=work_a).border = self.thin

                        HeaderBuilder.set_cell_value_safe(ws_area, row_area_temp, remain_a, remaining_area).font = self.remaining_font
                        ws_area.cell(row=row_area_temp, column=remain_a).fill = self.remaining_fill
                        ws_area.cell(row=row_area_temp, column=remain_a).alignment = self.center
                        ws_area.cell(row=row_area_temp, column=remain_a).border = self.thin

                        if is_total:
                            total_percent_c = percent_count
                            total_percent_a = percent_area
                            total_remain_c = remaining_count
                            total_remain_a = remaining_area

                        HeaderBuilder._apply_range(ws_count, row_count_temp, 1, row_count_temp, fact_c - 1, fill=group_fill)
                        HeaderBuilder._apply_range(ws_area,  row_area_temp,  1, row_area_temp,  fact_a - 1, fill=group_fill)

                        row_count_temp += 1
                        row_area_temp  += 1

                    op_end_row_area  = row_area_temp  - 1
                    op_end_row_count = row_count_temp - 1

                    HeaderBuilder.set_merged(ws_area,  row_area, 2, op_end_row_area,  2, op, font=self.bold, fill=self.operation_fills[op], align=self.left, border=self.thin)
                    HeaderBuilder.set_merged(ws_count, row_count, 2, op_end_row_count, 2, op, font=self.bold, fill=self.operation_fills[op], align=self.left, border=self.thin)

                    has_split = (split_by_itp or split_by_ogx)
                    if has_split and op_end_row_count > row_count:
                        self._merge_summary_cols_for_block(ws_count, row_count, op_end_row_count, percent_c, remain_c, total_percent_c, total_remain_c)
                        self._merge_summary_cols_for_block(ws_area,  row_area,  op_end_row_area,  percent_a, remain_a, total_percent_a, total_remain_a)

                    # ЖИРНАЯ горизонтальная линия в конце операции
                    draw_horizontal_divider(ws_area,  op_end_row_area, 1, remain_a, style="thick")
                    draw_horizontal_divider(ws_count, op_end_row_count, 1, remain_c, style="thick")

                    row_area  = row_area_temp + 1
                    row_count = row_count_temp + 1

                    prev_fact_count_by_group = current_fact_count_by_group
                    prev_fact_area_by_group  = current_fact_area_by_group

                if row_area - group_start_row_area > 1:
                    fill = PatternFill(start_color=self.group_colors.get(group_name, "FFFFFF"),
                                       end_color=self.group_colors.get(group_name, "FFFFFF"),
                                       fill_type="solid")
                    HeaderBuilder.set_merged(ws_area, group_start_row_area, 1, row_area - 1, 1, group_name,
                                             font=self.bold, align=self.center, border=self.thin, fill=fill)
                    HeaderBuilder.set_merged(ws_count, group_start_row_count, 1, row_count - 1, 1, group_name,
                                             font=self.bold, align=self.center, border=self.thin, fill=fill)

            for ws in [ws_area, ws_count]:
                self._auto_width(ws)
                self._unify_row_heights(ws, start_row=3, default_h=22)

            # под шапкой жирная горизонталь
            draw_horizontal_divider(ws_area,  header_bottom_a, 1, remain_a, style="thick")
            draw_horizontal_divider(ws_count, header_bottom_c, 1, remain_c, style="thick")

            # вертикальные разделители (тонкие)
            draw_vertical_divider(ws_area,  2, 1, ws_area.max_row, left=True, style="thin")
            draw_vertical_divider(ws_count, 2, 1, ws_count.max_row, left=True, style="thin")

            draw_vertical_divider(ws_area,  base_a, 1, ws_area.max_row, left=True, style="thin")
            draw_vertical_divider(ws_count, base_c, 1, ws_count.max_row, left=True, style="thin")

            for col in (fact_a, percent_a, work_a, remain_a):
                draw_vertical_divider(ws_area,  col, 1, ws_area.max_row, left=True, style="thin")
            for col in (fact_c, percent_c, work_c, remain_c):
                draw_vertical_divider(ws_count, col, 1, ws_count.max_row, left=True, style="thin")

            draw_outline(ws_area, 1, 1, ws_area.max_row, remain_a)
            draw_outline(ws_count, 1, 1, ws_count.max_row, remain_c)

            # Скрываем все недели, кроме последних 4
            self._hide_all_but_last_n_periods(ws_area, base_a, len(periods), keep_last=4)
            self._hide_all_but_last_n_periods(ws_count, base_c, len(periods), keep_last=4)

            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']

        open_file_crossplatform(out_path)

    #  МЕСЯЧНЫЙ
    def generate_monthly_combined(self, gz_year, plan_count, plan_area, split_by_ogx=False, split_by_itp=False):
        df = self.filter_by_gz_year(self.selection_filtered, gz_year)
        year_full = 2000 + int(gz_year)
        prev_year = year_full - 1
        next_year = year_full + 1
        months = ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь", "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]
        periods = [f"Декабрь {prev_year}"] + [f"{m} {year_full}" for m in months] + [f"Январь {next_year}", f"Февраль {next_year}"]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = f"месячный_отчет_{gz_year}_{timestamp}.xlsx"

        with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
            wb = writer.book
            ws_area = wb.create_sheet('Га.')
            ws_count = wb.create_sheet('Шт.')

            row_a, base_a, fact_a, percent_a, work_a, remain_a = self._build_header(ws_area, periods, include_itp=split_by_itp, include_ogx=split_by_ogx)
            row_c, base_c, fact_c, percent_c, work_c, remain_c = self._build_header(ws_count, periods, include_itp=split_by_itp, include_ogx=split_by_ogx)
            header_bottom_a = row_a - 1
            header_bottom_c = row_c - 1

            row_area = row_a
            row_count = row_c

            prev_fact_count_by_group = {}
            prev_fact_area_by_group = {}

            months_list = ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь", "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]

            for group_name, ops in self.operation_groups.items():
                group_start_row_area = row_area
                group_start_row_count = row_count

                for op in ops:
                    if op not in self.operation_map or self.operation_map[op] not in df.columns:
                        continue

                    for c in range(1, remain_a + 1):
                        _set_top_thick(ws_area, row_area, c)
                    for c in range(1, remain_c + 1):
                        _set_top_thick(ws_count, row_count, c)

                    # ЖИРНАЯ линия в НАЧАЛЕ операции (между row-1 и row)
                    draw_horizontal_divider(ws_area, row_area - 1, 1, remain_a, style="thick")
                    draw_horizontal_divider(ws_count, row_count - 1, 1, remain_c, style="thick")

                    col_name = self.operation_map[op]
                    fill = self.operation_fills[op]

                    # локальный флаг: разбивка по ИТП только для 1–3 операций
                    op_split_itp = self._op_split_itp(op, split_by_itp)
                    show_itp_col = op_split_itp

                    rows_full = list(self._iter_rows_for_operation(df, op, split_by_ogx, op_split_itp))


                    current_fact_count_by_group = {}
                    current_fact_area_by_group  = {}
                    for mask_func, labels in self._safe_iter(rows_full):
                        mask = mask_func()
                        fact_mask = mask & df[col_name].notna()
                        fact_count = int(fact_mask.sum())
                        fact_area  = truncate_float(float(df.loc[fact_mask, self.area_col].fillna(0).sum()), 1)
                        current_fact_count_by_group[tuple(labels)] = fact_count
                        current_fact_area_by_group[tuple(labels)]  = fact_area

                    rows = list(self._visible_rows_for_operation(rows_full, op, op_split_itp, split_by_ogx))
                    rows_safe = list(self._safe_iter(rows))

                    row_area_temp = row_area
                    row_count_temp = row_count
                    prev_itp = None

                    total_percent_c = total_percent_a = 0
                    total_remain_c = total_remain_a = 0

                    for mask_func, labels in self._safe_iter(rows):
                        mask = mask_func()
                        is_total = (len(labels) > 0 and labels[0] == "ВСЕГО")
                        write_col = 3
                        group_fill = PatternFill(start_color=self.group_colors.get(group_name, "FFFFFF"),
                                                 end_color=self.group_colors.get(group_name, "FFFFFF"),
                                                 fill_type="solid")

                        def put(ws, r, c, val, bold=False, filled=False, align_left=False):
                            cell = HeaderBuilder.set_cell_value_safe(ws, r, c, val)
                            cell.alignment = self.left if align_left else self.center
                            cell.font = self.bold if bold else self.normal
                            cell.border = self.thin
                            if filled:
                                cell.fill = self.subgroup_fill
                            return cell

                        double_split = (split_by_ogx and show_itp_col)

                        if double_split:
                            if len(labels) == 0:
                                # пустые метки — ничего не пишем
                                pass

                            elif len(labels) == 1 and labels[0] == "ВСЕГО":
                                put(ws_count, row_count_temp, write_col + 1, "ВСЕГО", bold=True, filled=True,
                                    align_left=True)
                                put(ws_area, row_area_temp, write_col + 1, "ВСЕГО", bold=True, filled=True,
                                    align_left=True)

                            elif len(labels) == 1:
                                if op == "1. Выдача нарядов":
                                    ogx_col = write_col + 1 if split_by_itp else write_col
                                    put(ws_count, row_count_temp, ogx_col, labels[0], align_left=True)
                                    put(ws_area, row_area_temp, ogx_col, labels[0], align_left=True)
                                    write_col = max(write_col, ogx_col) + 1
                                else:
                                    itp = labels[0]
                                    put(ws_count, row_count_temp, write_col, itp, bold=True, filled=True,
                                        align_left=True)
                                    put(ws_area, row_area_temp, write_col, itp, bold=True, filled=True, align_left=True)
                                    prev_itp = itp

                            elif len(labels) == 2:
                                itp, ogx = labels
                                if itp != prev_itp:
                                    put(ws_count, row_count_temp, write_col, itp, bold=True, filled=True,
                                        align_left=True)
                                    put(ws_area, row_area_temp, write_col, itp, bold=True, filled=True, align_left=True)
                                    prev_itp = itp
                                put(ws_count, row_count_temp, write_col + 1, ogx, align_left=True)
                                put(ws_area, row_area_temp, write_col + 1, ogx, align_left=True)

                            else:
                                # непредвидённая длина labels — безопасный фолбэк
                                put(ws_count, row_count_temp, write_col, "ВСЕГО", bold=True, filled=True,
                                    align_left=True)
                                put(ws_area, row_area_temp, write_col, "ВСЕГО", bold=True, filled=True, align_left=True)

                        else:
                            # корректная колонка для ОГХ
                            if split_by_ogx and split_by_itp and (op not in self.ops_with_itp_split):
                                ogx_col = write_col + 1
                            else:
                                ogx_col = write_col

                            if show_itp_col and len(labels) == 1 and labels[0] != "ВСЕГО":
                                lbl = labels[0]
                                itp_count = sum(1 for _, labs in rows_safe if (len(labs) > 0 and labs[0] == lbl))
                                end_row = row_count_temp + itp_count - 1
                                HeaderBuilder.set_merged(ws_count, row_count_temp, write_col, end_row, write_col, lbl,
                                                         font=self.bold, fill=self.subgroup_fill, align=self.center,
                                                         border=self.thin)
                                HeaderBuilder.set_merged(ws_area, row_area_temp, write_col, end_row, write_col, lbl,
                                                         font=self.bold, fill=self.subgroup_fill, align=self.center,
                                                         border=self.thin)
                                write_col += 1

                            if split_by_ogx:
                                lbl = labels[-1] if len(labels) > 1 else labels[0]
                                HeaderBuilder.set_merged(ws_count, row_count_temp, ogx_col, row_count_temp, ogx_col, lbl,
                                                         font=self.bold if lbl == "ВСЕГО" else self.normal,
                                                         fill=self.subgroup_fill if lbl == "ВСЕГО" else None,
                                                         align=self.center, border=self.thin)
                                HeaderBuilder.set_merged(ws_area, row_area_temp, ogx_col, row_area_temp, ogx_col, lbl,
                                                         font=self.bold if lbl == "ВСЕГО" else self.normal,
                                                         fill=self.subgroup_fill if lbl == "ВСЕГО" else None,
                                                         align=self.center, border=self.thin)
                                write_col = max(write_col, ogx_col) + 1
                            else:
                                lbl = labels[-1] if len(labels) > 1 else (labels[0] if labels else "")
                                HeaderBuilder.set_merged(ws_count, row_count_temp, ogx_col, row_count_temp, ogx_col,
                                                         lbl,
                                                         font=self.bold if lbl == "ВСЕГО" else self.normal,
                                                         fill=self.subgroup_fill if lbl == "ВСЕГО" else None,
                                                         align=self.center, border=self.thin)
                                HeaderBuilder.set_merged(ws_area, row_area_temp, ogx_col, row_area_temp, ogx_col, lbl,
                                                         font=self.bold if lbl == "ВСЕГО" else self.normal,
                                                         fill=self.subgroup_fill if lbl == "ВСЕГО" else None,
                                                         align=self.center, border=self.thin)
                                write_col = max(write_col, ogx_col) + 1

                        for j, period in enumerate(periods):
                            if period == f"Декабрь {prev_year}":
                                period_mask = (df[col_name].dt.year == prev_year) & (df[col_name].dt.month == 12)
                            elif period == f"Январь {next_year}":
                                period_mask = (df[col_name].dt.year == next_year) & (df[col_name].dt.month == 1)
                            elif period == f"Февраль {next_year}":
                                period_mask = (df[col_name].dt.year == next_year) & (df[col_name].dt.month == 2)
                            else:
                                m_name = period.split()[0]
                                m_idx = months_list.index(m_name) + 1
                                period_mask = (df[col_name].dt.year == year_full) & (df[col_name].dt.month == m_idx)

                            full_mask = mask & period_mask
                            cnt = int(full_mask.sum())
                            area = float(df.loc[full_mask, self.area_col].fillna(0).sum())

                            cell_cnt = HeaderBuilder.set_cell_value_safe(ws_count, row_count_temp, base_c + j, cnt)
                            cell_cnt.alignment = self.center; cell_cnt.border = self.thin; cell_cnt.fill = fill
                            if is_total: cell_cnt.font = self.bold

                            cell_area = HeaderBuilder.set_cell_value_safe(ws_area, row_area_temp, base_a + j, truncate_float(area, 1))
                            cell_area.alignment = self.center; cell_area.border = self.thin; cell_area.fill = fill
                            if is_total: cell_area.font = self.bold

                        key = tuple(labels)
                        fact_count = current_fact_count_by_group.get(key, 0)
                        fact_area  = current_fact_area_by_group.get(key, 0.0)

                        if is_total:
                            percent_count = round((fact_count / plan_count * 100), 1) if plan_count > 0 else 0
                            percent_area  = round((fact_area  / plan_area  * 100), 1) if plan_area  > 0 else 0
                        else:
                            percent_count = round((fact_count / plan_count * 100), 2) if plan_count > 0 else 0
                            percent_area  = round((fact_area  / plan_area  * 100), 2) if plan_area  > 0 else 0

                        work_count = "-"
                        work_area  = "-"
                        if op != "1. Выдача нарядов":
                            prev_count = self._prev_total_for_labels(prev_fact_count_by_group, labels)
                            prev_area  = self._prev_total_for_labels(prev_fact_area_by_group,  labels)
                            work_count = max(prev_count - fact_count, 0)
                            work_area  = truncate_float(max(prev_area - fact_area, 0.0), 1)

                        remaining_count = max(plan_count - fact_count, 0)
                        remaining_area  = truncate_float(max(plan_area  - fact_area,  0.0), 1)

                        HeaderBuilder.set_cell_value_safe(ws_count, row_count_temp, fact_c, fact_count).font = self.bold
                        ws_count.cell(row=row_count_temp, column=fact_c).fill = self.fact_fill
                        ws_count.cell(row=row_count_temp, column=fact_c).alignment = self.center
                        ws_count.cell(row=row_count_temp, column=fact_c).border = self.thin

                        HeaderBuilder.set_cell_value_safe(ws_count, row_count_temp, percent_c, f"{percent_count}%").font = self.bold
                        ws_count.cell(row=row_count_temp, column=percent_c).fill = self.percent_fill
                        ws_count.cell(row=row_count_temp, column=percent_c).alignment = self.center
                        ws_count.cell(row=row_count_temp, column=percent_c).border = self.thin

                        HeaderBuilder.set_cell_value_safe(ws_count, row_count_temp, work_c, work_count).font = self.bold
                        ws_count.cell(row=row_count_temp, column=work_c).fill = self.work_fill
                        ws_count.cell(row=row_count_temp, column=work_c).alignment = self.center
                        ws_count.cell(row=row_count_temp, column=work_c).border = self.thin

                        HeaderBuilder.set_cell_value_safe(ws_count, row_count_temp, remain_c, remaining_count).font = self.remaining_font
                        ws_count.cell(row=row_count_temp, column=remain_c).fill = self.remaining_fill
                        ws_count.cell(row=row_count_temp, column=remain_c).alignment = self.center
                        ws_count.cell(row=row_count_temp, column=remain_c).border = self.thin

                        HeaderBuilder.set_cell_value_safe(ws_area, row=row_area_temp, col=fact_a, value=fact_area).font = self.bold
                        ws_area.cell(row=row_area_temp, column=fact_a).fill = self.fact_fill
                        ws_area.cell(row=row_area_temp, column=fact_a).alignment = self.center
                        ws_area.cell(row=row_area_temp, column=fact_a).border = self.thin

                        HeaderBuilder.set_cell_value_safe(ws_area, row=row_area_temp, col=percent_a, value=f"{percent_area}%").font = self.bold
                        ws_area.cell(row=row_area_temp, column=percent_a).fill = self.percent_fill
                        ws_area.cell(row=row_area_temp, column=percent_a).alignment = self.center
                        ws_area.cell(row=row_area_temp, column=percent_a).border = self.thin

                        HeaderBuilder.set_cell_value_safe(ws_area, row=row_area_temp, col=work_a, value=work_area).font = self.bold
                        ws_area.cell(row=row_area_temp, column=work_a).fill = self.work_fill
                        ws_area.cell(row=row_area_temp, column=work_a).alignment = self.center
                        ws_area.cell(row=row_area_temp, column=work_a).border = self.thin

                        HeaderBuilder.set_cell_value_safe(ws_area, row=row_area_temp, col=remain_a, value=remaining_area).font = self.remaining_font
                        ws_area.cell(row=row_area_temp, column=remain_a).fill = self.remaining_fill
                        ws_area.cell(row=row_area_temp, column=remain_a).alignment = self.center
                        ws_area.cell(row=row_area_temp, column=remain_a).border = self.thin

                        if is_total:
                            total_percent_c = percent_count
                            total_percent_a = percent_area
                            total_remain_c = remaining_count
                            total_remain_a = remaining_area

                        HeaderBuilder._apply_range(ws_count, row_count_temp, 1, row_count_temp, fact_c - 1, fill=group_fill)
                        HeaderBuilder._apply_range(ws_area,  row_area_temp,  1, row_area_temp,  fact_a - 1, fill=group_fill)

                        row_count_temp += 1
                        row_area_temp  += 1

                    op_end_row_area  = row_area_temp  - 1
                    op_end_row_count = row_count_temp - 1
                    HeaderBuilder.set_merged(ws_area,  row_area, 2, op_end_row_area,  2, op, font=self.bold, fill=self.operation_fills[op], align=self.left, border=self.thin)
                    HeaderBuilder.set_merged(ws_count, row_count, 2, op_end_row_count, 2, op, font=self.bold, fill=self.operation_fills[op], align=self.left, border=self.thin)

                    has_split = (split_by_itp or split_by_ogx)
                    if has_split and op_end_row_count > row_count:
                        self._merge_summary_cols_for_block(ws_count, row_count, op_end_row_count, percent_c, remain_c, total_percent_c, total_remain_c)
                        self._merge_summary_cols_for_block(ws_area,  row_area,  op_end_row_area,  percent_a, remain_a, total_percent_a, total_remain_a)

                    # ЖИРНАЯ горизонтальная линия в конце операции
                    draw_horizontal_divider(ws_area,  op_end_row_area, 1, remain_a, style="thick")
                    draw_horizontal_divider(ws_count, op_end_row_count, 1, remain_c, style="thick")

                    row_area  = row_area_temp + 1
                    row_count = row_count_temp + 1

                    prev_fact_count_by_group = current_fact_count_by_group
                    prev_fact_area_by_group  = current_fact_area_by_group

                if row_area - group_start_row_area > 1:
                    fill = PatternFill(start_color=self.group_colors.get(group_name, "FFFFFF"),
                                       end_color=self.group_colors.get(group_name, "FFFFFF"),
                                       fill_type="solid")
                    HeaderBuilder.set_merged(ws_area, group_start_row_area, 1, row_area - 1, 1, group_name,
                                             font=self.bold, align=self.center, border=self.thin, fill=fill)
                    HeaderBuilder.set_merged(ws_count, group_start_row_count, 1, row_count - 1, 1, group_name,
                                             font=self.bold, align=self.center, border=self.thin, fill=fill)

            for ws in [ws_area, ws_count]:
                self._auto_width(ws)
                self._unify_row_heights(ws, start_row=3, default_h=22)

            # под шапкой тонкая горизонталь
            draw_horizontal_divider(ws_area,  header_bottom_a, 1, remain_a, style="thick")
            draw_horizontal_divider(ws_count, header_bottom_c, 1, remain_c, style="thick")

            # вертикальные разделители (тонкие)
            draw_vertical_divider(ws_area,  2, 1, ws_area.max_row, left=True, style="thin")
            draw_vertical_divider(ws_count, 2, 1, ws_count.max_row, left=True, style="thin")

            draw_vertical_divider(ws_area,  base_a, 1, ws_area.max_row, left=True, style="thin")
            draw_vertical_divider(ws_count, base_c, 1, ws_count.max_row, left=True, style="thin")

            for col in (fact_a, percent_a, work_a, remain_a):
                draw_vertical_divider(ws_area,  col, 1, ws_area.max_row, left=True, style="thin")
            for col in (fact_c, percent_c, work_c, remain_c):
                draw_vertical_divider(ws_count, col, 1, ws_count.max_row, left=True, style="thin")

            draw_outline(ws_area, 1, 1, ws_area.max_row, remain_a)
            draw_outline(ws_count, 1, 1, ws_count.max_row, remain_c)

            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']

        open_file_crossplatform(out_path)
